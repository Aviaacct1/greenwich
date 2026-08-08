"""Generator increment: the Macro block and real-to-nominal (queue item 4, note 17).

Blueprint 06 lists a Macro block that the skeleton never grew back after v0. This
restores it as an additive chain step (same shape as capex_block and
financing_group): it adds a Macro sheet holding CPI and construction cost
inflation with cumulative indices from the base year, then adds nominal
presentation from the existing REAL calc rows, so nothing already wired is
shifted. CPI drives the P&L real-to-nominal; the construction index drives Capex
real-to-nominal, exactly the two hooks macro_feed was built to supply.

Inputs are the workbook so far, the project header (start_year, last_actual_year,
model_term_years), and a macro T1 file (macro_feed output: cpi and
construction_index rows, annual percentages, segment carrying the source). The
base year is the last actual year, so the index is 1.00 in base-year money and
forecast years inflate from there. Where a series runs short of the model term
its last value is held flat, stated on the sheet. Author: Avia Solutions."""
import sys, openpyxl
from openpyxl.styles import Font

ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
FIRST_YCOL = 12
T1M = ["metric_code", "segment", "case_id", "year", "value", "unit",
       "temporality", "driver_type", "step_date", "step_value", "repeat_years", "source"]


def read_header(path):
    h = {}
    for line in open(path, encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"):
            continue
        k, v = line.rstrip("\n").split("\t")
        h[k] = v
    return h


def read_macro(path):
    """Return {metric: {year: pct}} and {metric: source}, first source per metric."""
    series, src = {}, {}
    for line in open(path, encoding="utf-8"):
        if line.startswith("#") or not line.strip():
            continue
        r = dict(zip(T1M, line.rstrip("\n").split("\t")))
        m = r["metric_code"]
        if m not in ("cpi", "construction_index"):
            continue
        series.setdefault(m, {})[int(r["year"])] = float(r["value"])
        src.setdefault(m, r["source"])
    return series, src


def col(j):
    return openpyxl.utils.get_column_letter(FIRST_YCOL + j)


def main(wb_in, header_path, macro_path, wb_out):
    H = read_header(header_path)
    Y0 = int(H["start_year"])
    N = int(H["model_term_years"])
    base = int(H.get("last_actual_year", Y0 - 1))
    years = [Y0 + j for j in range(N)]
    series, src = read_macro(macro_path)

    def filled(metric):
        """Values across the term, holding the last available value flat; note the cut year."""
        s = series.get(metric, {})
        if not s:
            return [None] * N, None
        out, last, held_from = [], None, None
        for y in years:
            if y in s:
                last = s[y]
            elif last is not None and held_from is None:
                held_from = y
            out.append(last)
        return out, held_from

    wb = openpyxl.load_workbook(wb_in)
    if "Macro" in wb.sheetnames:
        del wb["Macro"]
    ws = wb.create_sheet("Macro")
    ws["A1"] = "=Control!B1"; ws["A1"].font = BOLD
    ws["A3"] = f"Macro block: base year {base} = index 1.00; forecast years inflate from the base."
    ws["A3"].font = ARIAL
    for j in range(N):
        ws.cell(row=6, column=FIRST_YCOL + j, value=years[j]).font = BOLD

    # (metric, pct_row, index_row, pct_label, index_label)
    specs = [("cpi", 8, 9, "CPI inflation [%]", f"CPI cumulative index (base {base} = 1.00)"),
             ("construction_index", 11, 12, "Construction cost inflation [%]",
              f"Construction cumulative index (base {base} = 1.00)")]
    notes = []
    index_row_of = {}
    # base year column: index is 1.00 there, compounds forward after, deflates backward before.
    # bcol == -1 means the base is the year just before the spine (the common case).
    bcol = base - Y0
    if bcol < -1:
        notes.append(f"Base year {base} precedes the spine start {Y0} by more than one year; inflation "
                     f"for the intervening years is not supplied, so the index anchors at {Y0}.")
    for metric, pr, ir, plab, ilab in specs:
        vals, held = filled(metric)
        ws.cell(row=pr, column=4, value=plab).font = ARIAL
        ws.cell(row=ir, column=4, value=ilab).font = BOLD
        for j in range(N):
            c = col(j)
            v = vals[j]
            ws[f"{c}{pr}"] = ("" if v is None else v); ws[f"{c}{pr}"].font = ARIAL
            # cumulative index anchored at the base year = 1.00 (not at the first spine column)
            if v is None:
                ws[f"{c}{ir}"] = 1
            elif j == bcol:
                ws[f"{c}{ir}"] = 1                                  # base year
            elif j > bcol:
                prev = f"{col(j-1)}{ir}" if j - 1 >= 0 else "1"     # forward: compound from the base
                ws[f"{c}{ir}"] = f"={prev}*(1+{c}{pr}/100)"
            else:
                nxt = col(j + 1)                                    # backward: deflate toward the base
                ws[f"{c}{ir}"] = f"={nxt}{ir}/(1+{nxt}{pr}/100)"
            ws[f"{c}{ir}"].font = ARIAL
        index_row_of[metric] = ir
        s = src.get(metric, "None")
        ws.cell(row=ir + 1, column=4, value=f"Source: {s}").font = ARIAL
        if held:
            notes.append(f"{plab} held flat at its last value from {held} (series runs short of the term).")
    for k, n in enumerate(notes):
        ws.cell(row=16 + k, column=4, value="Note: " + n).font = ARIAL
    if not notes:
        ws.cell(row=16, column=4, value="Note: both series cover the full model term.").font = ARIAL

    cpi_ir = index_row_of.get("cpi")
    con_ir = index_row_of.get("construction_index")

    # Nominal presentation on Operations Summary (CPI): real x cumulative CPI index
    if "Operations Summary" in wb.sheetnames and cpi_ir:
        osw = wb["Operations Summary"]
        osw.cell(row=30, column=4,
                 value="Nominal presentation (real x Macro CPI cumulative index)").font = BOLD
        osw.cell(row=31, column=4, value="Total revenue (nominal)").font = ARIAL
        osw.cell(row=32, column=4, value="EBITDA (nominal)").font = BOLD
        for j in range(N):
            c = col(j)
            osw[f"{c}31"] = f"={c}10*Macro!{c}{cpi_ir}"; osw[f"{c}31"].font = ARIAL
            osw[f"{c}32"] = f"={c}12*Macro!{c}{cpi_ir}"; osw[f"{c}32"].font = ARIAL

    # Nominal Capex (construction index): real growth and maintenance capex x cumulative index.
    # Rows discovered by label (as financing_group does), not hard-coded, so a change to the
    # Capex layout does not silently point these at the wrong rows.
    def find_row(ws, label, coln=4):
        for r in ws.iter_rows(min_col=coln, max_col=coln):
            if r[0].value == label:
                return r[0].row
        return None
    if "Capex" in wb.sheetnames and con_ir:
        cxw = wb["Capex"]
        gc = find_row(cxw, "Total expansion (growth) capex")
        mt = find_row(cxw, "Maintenance capex")
        if gc and mt:
            cxw.cell(row=32, column=4,
                     value="Nominal presentation (real x Macro construction cumulative index)").font = BOLD
            cxw.cell(row=33, column=4, value="Total growth capex (nominal)").font = ARIAL
            cxw.cell(row=34, column=4, value="Maintenance capex (nominal)").font = ARIAL
            for j in range(N):
                c = col(j)
                cxw[f"{c}33"] = f"={c}{gc}*Macro!{c}{con_ir}"; cxw[f"{c}33"].font = ARIAL
                cxw[f"{c}34"] = f"={c}{mt}*Macro!{c}{con_ir}"; cxw[f"{c}34"].font = ARIAL
        else:
            ws.cell(row=20, column=4,
                    value="Note: Capex nominal rows skipped, growth or maintenance capex row not found by label.").font = ARIAL

    # R110 nominal identity tripwire: recompute nominal independently and compare on the Macro sheet
    if cpi_ir and "Operations Summary" in wb.sheetnames:
        c0, cZ = col(0), col(N - 1)
        ws.cell(row=22, column=4,
                value="Nominal identity check (Operations Summary nominal revenue = real x CPI index)").font = BOLD
        ws.cell(row=22, column=7,
                value=(f"=IF(SUMPRODUCT(--(ABS('Operations Summary'!{c0}31:{cZ}31-"
                       f"'Operations Summary'!{c0}10:{cZ}10*{c0}{cpi_ir}:{cZ}{cpi_ir})>0.000001))=0,"
                       f"\"Ok\",\"ERROR\")")).font = BOLD

    wb.calculation.fullCalcOnLoad = True
    wb.properties.creator = "Avia Solutions"
    wb.properties.lastModifiedBy = "Avia Solutions"
    wb.save(wb_out)
    have = [m for m in ("cpi", "construction_index") if m in series]
    print(f"saved {wb_out} | Macro sheet added | series: {', '.join(have) or 'none'} | "
          f"base year {base} | nominal rows on Operations Summary and Capex")


if __name__ == "__main__":
    # usage: macro_block.py <wb_in> <header> <macro_t1> <wb_out>
    main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
