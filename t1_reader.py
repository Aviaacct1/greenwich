"""T1 reader v3 (generator v1). v2 corrected the year anchoring; v3 adds the
two behaviours the Studio's 4c interaction model requires of the same table:
(a) EVENT LISTS: multiple one_off_step rows per line accumulate into the uplift
row, combining multiplicatively where two events share a year (v2 let a second
row zero the first's year; repeat_years still expands for back-compat but the
Studio writes discrete rows); (b) MULTI-ANCHOR LEVELS: several level rows per
line (the base seed plus any manual grid cells) each anchor their own year as a
constant, with chain formulas only between anchors, so a manual cell feeds the
chain exactly as in Excel. Cells before the first anchor stay blank, never
fabricated. Spine from project_header.tsv. Author: Avia Solutions."""
import sys, re, openpyxl
from openpyxl.utils import get_column_letter
from t1_maps import build_maps

FC = 12

def read_header(p):
    h={}
    for line in open(p,encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"): continue
        k,v=line.rstrip("\n").split("\t"); h[k]=v
    return h

def load_t1(path):
    rows=[]
    with open(path,encoding="utf-8") as fh:
        for line in fh:
            if not line.strip() or line.startswith("#"): continue
            p=line.rstrip("\n").split("\t")
            rows.append(dict(zip(["metric_code","segment","case_id","year","value","unit",
                "temporality","driver_type","step_date","step_value","repeat_years","source"],p)))
    return rows

def calc_row_for(wb,lab):
    """Calc-sheet cell to re-anchor for a level label: non-aero category revenue
    row or opex line row. Traffic and aero rate rows read Inputs every year and
    need no rewrite."""
    if lab.endswith(" - base year revenue"):
        cat=lab[:-len(" - base year revenue")]
        for sn in wb.sheetnames:
            if sn=="Non-aero" or sn.startswith("Non Aero"):
                ws=wb[sn]
                for row in ws.iter_rows(min_col=3,max_col=3):
                    if row[0].value==cat:
                        return ws,row[0].row+9      # rev row per block grammar
    if lab.endswith(" - base year"):
        cat=lab[:-len(" - base year")]
        for sn in wb.sheetnames:
            if sn=="Opex" or sn.startswith("Opex "):
                ws=wb[sn]
                for row in ws.iter_rows(min_col=4,max_col=4):
                    if row[0].value==cat:
                        return ws,row[0].row
    return None,None

def populate(skel_in,t1_path,header_path,line_sets_path,tier,out_path):
    H=read_header(header_path)
    Y0=int(H["start_year"]); NY=int(H["model_term_years"])
    LA=int(H.get("last_actual_year",Y0-1)); jLA=LA-Y0
    YC=[get_column_letter(FC+i) for i in range(NY)]
    wb=openpyxl.load_workbook(skel_in)
    inp,ctrl=wb["Inputs"],wb["Control"]
    lbl_row={}
    for r in range(30,600):
        v=ctrl.cell(row=r,column=7).value
        if isinstance(v,str): lbl_row[v]=r
    sel={}
    for row in inp.iter_rows():
        c=row[5]
        if isinstance(c.value,str) and c.value.startswith("=F") and row[8].value:
            m=re.match(r"=Control!\$G\$(\d+)",str(inp.cell(row=c.row-6,column=4).value or ""))
            if m: sel[int(m.group(1))]=c.row
    BY_KEY,_,_=build_maps(line_sets_path,int(tier))
    def label_for(r):
        return BY_KEY.get((r["metric_code"],r["segment"],
            "per_pax" if r["driver_type"]=="per_pax" else r["driver_type"]))
    t1=load_t1(t1_path)
    growth={r["metric_code"]:float(r["value"]) for r in t1 if r["driver_type"]=="overlay"}
    # group by label
    levels,steps,flats,srcs={},{},{},{}
    for r in t1:
        lab=label_for(r)
        if not lab or lab not in lbl_row or lbl_row[lab] not in sel: continue
        srcs.setdefault(lab,r["source"])
        d=r["driver_type"]
        if d=="one_off_step":
            steps.setdefault(lab,[]).append((int(r["step_date"]),float(r["step_value"]),
                                             int(r["repeat_years"] or 0)))
        elif d in ("level","per_pax") and r["year"]!="ALL":
            g=growth.get(r["metric_code"],0.0) if r["metric_code"].startswith(("pax_","atm_")) else 0.0
            levels.setdefault(lab,[]).append((int(r["year"])-Y0,float(r["value"]),g))
        elif d in ("elasticity_pax","elasticity_sqm"):
            flats[lab]=float(r["value"])
    n=0
    for lab,anchors in levels.items():
        br=sel[lbl_row[lab]]-5
        anchors=sorted(a for a in anchors if 0<=a[0]<NY)
        if not anchors: continue
        firsts=[a[0] for a in anchors]
        for idx,(j0,v,g) in enumerate(anchors):
            nxt=anchors[idx+1][0] if idx+1<len(anchors) else NY
            inp.cell(row=br,column=FC+j0,value=v)
            for j in range(j0+1,nxt):
                inp.cell(row=br,column=FC+j,value=f"={YC[j-1]}{br}*(1+{g})")
        for j in range(0,firsts[0]):
            inp.cell(row=br,column=FC+j).value=None
        inp.cell(row=br,column=48,value=f"Source: {srcs[lab]}")
        # forecast anchors beyond the block seed re-anchor the CALC row too
        # (grid disposes: a manual cell must feed the P&L chain, not just Inputs)
        late=[j0 for j0,_,_ in anchors if j0>max(0,jLA) and j0!=firsts[0]]
        if late:
            ws,cr=calc_row_for(wb,lab)
            if ws is not None:
                srow=sel[lbl_row[lab]]
                for j0 in late:
                    ws.cell(row=cr,column=FC+j0,value=f"=Inputs!{YC[j0]}{srow}")
        n+=1
    for lab,evs in steps.items():
        br=sel[lbl_row[lab]]-5
        ae=[y1 for y1,_,_ in evs if y1<=LA]   # anchors dated in the actual era
        for j in range(NY):
            yr=Y0+j
            f=1.0
            if yr>LA:                          # events never modify a reported actual year
                for y1,step,cyc in evs:
                    hit=(yr==y1) or (cyc and yr>y1 and (yr-y1)%cyc==0)
                    if hit: f*=(1.0+step)
            inp.cell(row=br,column=FC+j,value=round(f-1.0,10))
        note=srcs[lab]+(f"; NOTE: event anchor(s) in the actual era ({','.join(str(y) for y in ae)}) "
                        f"suppressed in actual years, forecast recurrences still apply" if ae else "")
        inp.cell(row=br,column=48,value=f"Source: {note}")
        n+=1
    for lab,v in flats.items():
        br=sel[lbl_row[lab]]-5
        inp.cell(row=br,column=FC,value=v)
        for j in range(1,NY): inp.cell(row=br,column=FC+j,value=f"={YC[j-1]}{br}")
        inp.cell(row=br,column=48,value=f"Source: {srcs[lab]}")
        n+=1
    wb["Cover"]["C9"]="Inputs populated from T1 assumptions table; year-anchored rows on the header spine; event lists accumulated; multi-anchor levels chain between anchors; sources per row on Inputs col AV"
    wb.properties.creator=wb.properties.lastModifiedBy="Avia Solutions"
    wb.save(out_path)
    print("T1 lines applied:",n,"| spine",Y0,"to",Y0+NY-1)
    return n

if __name__=="__main__":
    populate(sys.argv[1],sys.argv[2],sys.argv[3],sys.argv[4],sys.argv[5],sys.argv[6])
