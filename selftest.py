"""Regression self-test suite for the Avia model generator (17 July 2026).

One command that rebuilds the demos and asserts the key numbers and every coherence
check, so any correction lands with instant re-verification. It exercises the whole
chain and the session's increments: tier-1 build, macro block real-to-nominal, aero
and opex events, aero rate calibration, and the refresh automation. Numbers are
recalculated with LibreOffice (one headless pass over all workbooks). Exits non-zero
on any failure, and prints a PASS/FAIL line per assertion.

Usage: python3 selftest.py <model_refs_dir> <work_dir> [mode]
  mode = all (default) | build (builds + tools, no recalc) | verify (recalc + assert)
The build/verify split exists only so each phase fits a short shell timeout; a machine
with no such limit runs 'all'. build and verify share the work dir.
Author: Avia Solutions."""
import sys, os, shutil, subprocess, glob
import config
import openpyxl
from openpyxl.utils import get_column_letter as gl

FAILS = []
def check(cond, msg):
    print(("PASS" if cond else "FAIL") + ": " + msg)
    if not cond:
        FAILS.append(msg)

def run(*args):
    r = subprocess.run([config.python_exe(), *args], capture_output=True, text=True)
    if r.returncode != 0:
        print(r.stdout.strip()); print(r.stderr.strip())
        raise SystemExit("self-test: command failed: " + " ".join(str(a) for a in args))
    return r.stdout.strip()

def cy(y): return 12 + (y - 2024)

def build_phase(ref, wd):
    fx = os.path.join(ref, "selftest_fixtures")
    if os.path.isdir(wd): shutil.rmtree(wd, ignore_errors=True)
    os.makedirs(wd, exist_ok=True)
    for f in glob.glob(os.path.join(ref, "*.py")) + [os.path.join(ref, "line_sets.tsv")]:
        dst = os.path.join(wd, os.path.basename(f))
        shutil.copy(f, dst)
        try: os.chmod(dst, 0o644)
        except OSError: pass
    hdr = os.path.join(fx, "header.tsv"); act = os.path.join(fx, "actuals.tsv")
    s = lambda n: os.path.join(wd, n)

    # ---- builds (openpyxl, fast) ----
    run(s("build_model.py"), wd, hdr, os.path.join(fx, "t1_base.tsv"),   act, s("base.xlsx"))
    run(s("build_model.py"), wd, hdr, os.path.join(fx, "t1_events.tsv"), act, s("evt.xlsx"))
    run(s("build_model.py"), wd, hdr, os.path.join(fx, "t1_base.tsv"),   act, s("macro.xlsx"),
        "-", "-", "-", "-", "1", os.path.join(fx, "t1_macro.tsv"))
    # end-to-end aero calibration: calibrated rate + matching pax should reproduce ingested revenue
    run(s("build_model.py"), wd, os.path.join(fx, "e2e_header.tsv"), os.path.join(fx, "e2e_t1.tsv"),
        os.path.join(fx, "e2e_act.tsv"), s("e2e.xlsx"))
    # adversarial: macro file carrying history, base year inside the spine (C1 regression guard)
    run(s("macro_block.py"), s("base.xlsx"), os.path.join(fx, "header_hist.tsv"),
        os.path.join(fx, "t1_macro_hist.tsv"), s("macrohist.xlsx"))

    # ---- aero rate calibration ----
    out = run(s("calibrate_aero.py"), os.path.join(fx, "t6_confirmed.tsv"),
              os.path.join(fx, "t1_vendor_actuals.tsv"), os.path.join(fx, "t1_traffic.tsv"),
              s("line_sets.tsv"), s("calib"))
    check("identity checks ALL PASS" in out, "aero calibration identity checks pass")
    # adversarial: split aero line, revenue must be apportioned by share not dropped
    run(s("calibrate_aero.py"), os.path.join(fx, "t6_aero_split.tsv"),
        os.path.join(fx, "t1_vendor_aero_split.tsv"), os.path.join(fx, "t1_traffic_split.tsv"),
        s("line_sets.tsv"), s("calibsplit"))
    rates = open(os.path.join(s("calibsplit"), "t1_aero_rates.tsv")).read()
    check("rev_psc\ttotal\t1\t2023\t4.167273" in rates, "aero split: PSC share-attributed 0.6 x 38.2 over pax")
    check("rev_security\ttotal\t1\t2023\t2.778182" in rates, "aero split: security share-attributed 0.4 x 38.2")

    # ---- refresh automation ----
    est = s("est"); os.makedirs(est)
    open(os.path.join(est, "weo_oct2026.csv"), "w").write("x")
    shutil.copy(os.path.join(ref, "macro_watchlist.tsv"), s("wl.tsv"))
    run(s("refresh_delta.py"), est, s("mf.tsv"), s("wl.tsv"), s("delta.md"))
    dr = open(s("delta.md")).read()
    check("IMF WEO: vintage oct2026" in dr, "refresh_delta leads with the new IMF WEO vintage")
    run(s("macro_adopt.py"), os.path.join(fx, "t1_macro.tsv"),
        os.path.join(fx, "t1_macro_newvintage.tsv"), "Selftest", s("adopt.md"))
    check("+0.3000" in open(s("adopt.md")).read(), "macro_adopt shows the +0.3 vintage movement")

def verify_phase(ref, wd):
    s = lambda n: os.path.join(wd, n)
    # ---- recalc base/evt/macro in one LibreOffice pass ----
    rc = os.path.join(wd, "rc"); os.makedirs(rc, exist_ok=True)
    subprocess.run(["libreoffice", "--headless", "--calc", "--convert-to", "xlsx", "--outdir", rc,
                    s("base.xlsx"), s("evt.xlsx"), s("macro.xlsx"), s("e2e.xlsx"), s("macrohist.xlsx")],
                   capture_output=True, timeout=240)
    B = openpyxl.load_workbook(os.path.join(rc, "base.xlsx"), data_only=True)
    E = openpyxl.load_workbook(os.path.join(rc, "evt.xlsx"), data_only=True)
    M = openpyxl.load_workbook(os.path.join(rc, "macro.xlsx"), data_only=True)
    A2 = openpyxl.load_workbook(os.path.join(rc, "e2e.xlsx"), data_only=True)
    MH = openpyxl.load_workbook(os.path.join(rc, "macrohist.xlsx"), data_only=True)

    def os_cell(wb, r, y): return wb["Operations Summary"].cell(row=r, column=cy(y)).value
    def checks_ok(wb):
        ck = wb["Checks"]
        st = {ck.cell(row=r, column=2).value: ck.cell(row=r, column=7).value for r in range(5, 25)
              if ck.cell(row=r, column=2).value}
        return st.get("Overall model status") == "Ok"

    # base numbers (pinned regression values)
    check(abs(os_cell(B, 10, 2024) - 229.525) < 1e-2, "base 2024 revenue = 229.525")
    check(abs(os_cell(B, 12, 2024) - 194.675) < 1e-2, "base 2024 EBITDA = 194.675")
    check(checks_ok(B), "base: all coherence checks Ok")

    # events fire (ratio vs base)
    def row_for(wb, sheet, label):
        for r in wb[sheet].iter_rows(min_col=4, max_col=4):
            if r[0].value == label: return r[0].row
    psc = row_for(B, "Aero", "Passenger service charge")
    stf = row_for(B, "Opex", "Staff costs")
    def rat(wb, sheet, row, y): return wb[sheet].cell(row=row, column=cy(y)).value / B[sheet].cell(row=row, column=cy(y)).value
    check(abs(rat(E, "Aero", psc, 2028) - 1.0) < 1e-4, "event: PSC unchanged through 2028")
    check(abs(rat(E, "Aero", psc, 2029) - 1.05) < 1e-4, "event: PSC +5% reset from 2029")
    check(abs(rat(E, "Opex", stf, 2030) - 0.90) < 1e-4, "event: staff -10% step from 2030")
    check(checks_ok(E), "events: all coherence checks Ok")

    # macro real-to-nominal
    mac = M["Macro"]
    cpi25 = mac.cell(row=9, column=cy(2025)).value
    check(abs(cpi25 - 1.06502) < 1e-4, "macro: CPI cumulative index 2025 = 1.06502")
    realrev = os_cell(M, 10, 2024); nomrev = os_cell(M, 31, 2024); idx = mac.cell(row=9, column=cy(2024)).value
    check(abs(nomrev - realrev * idx) < 1e-6, "macro: nominal revenue = real x CPI index")

    # aero calibration end to end: calibrated PSC rate x matching pax reproduces the ingested revenue
    ae = A2["Aero"]
    psc_row = row_for(A2, "Aero", "Passenger service charge")
    psc_e2e = ae.cell(row=psc_row, column=cy(2024)).value
    check(abs(psc_e2e - 41.5) < 1e-2, "aero calibration end to end: calibrated PSC rate reproduces ingested 41.5 in the workbook")

    # C1 guard: macro index anchored at the base year when the base sits inside the spine
    mh = MH["Macro"]
    i_base = mh.cell(row=9, column=cy(2027)).value
    mh_real = MH["Operations Summary"].cell(row=10, column=cy(2027)).value
    mh_nom = MH["Operations Summary"].cell(row=31, column=cy(2027)).value
    check(abs(i_base - 1.0) < 1e-9, "macro history: CPI index = 1.00 at the base year inside the spine (C1)")
    check(abs(mh_nom - mh_real) < 1e-6, "macro history: nominal = real at the base year (C1)")

def summarize():
    print()
    if FAILS:
        print(f"SELF-TEST: {len(FAILS)} assertion(s) FAILED")
        for m in FAILS: print("  -", m)
        raise SystemExit(1)
    print("SELF-TEST: all assertions in this phase green")

def main(ref, wd, mode="all"):
    if mode in ("all", "build"): build_phase(ref, wd)
    if mode in ("all", "verify"): verify_phase(ref, wd)
    summarize()

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else "all")
