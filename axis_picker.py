"""Generator v1 increment 4a: per-axis scenario picker (Scanner pattern,
blueprint 06 section 8 design consequence 2). Adds an axis picker block to
Control (independent case choice per Traffic / Aero / Non-aero / Opex axis);
each line's case-choice cell becomes an inherit-from-axis formula with a
per-line override column (blank = inherit, number = override). The Inputs
INDEX selector grammar is untouched: calc sheets still read only selector
rows. Lines are classified to axes by their Control label against the same
category lists the skeleton generator uses. Author: Avia Solutions."""
import sys, openpyxl
from openpyxl.styles import Font

ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)

TRAFFIC_LABELS = {"Passengers - domestic", "Passengers - international",
                  "ATMs - domestic", "ATMs - international", "ATMs - cargo", "ATMs - other"}
NONAERO_CATS = ["Duty free", "Specialty retail", "Food & beverage", "Advertising",
                "Car parking", "Car rental", "Lounge", "Property rental",
                "Fuel throughput", "Other non-aero"]
OPEX_CATS = ["Staff costs", "Utilities", "Repairs and maintenance", "Insurance",
             "Rent and rates", "Marketing", "Cleaning", "Other opex", "SPV / corporate costs"]

AXIS_ROW = {"Traffic": 9, "Aero": 10, "Non-aero": 11, "Opex": 12}

def axis_of(label):
    if label in TRAFFIC_LABELS:
        return "Traffic"
    if label.endswith("unit rate") or label.endswith("reset uplift"):
        return "Aero"
    stem = label.split(" - ")[0]
    if stem in NONAERO_CATS:
        return "Non-aero"
    if stem in OPEX_CATS:
        return "Opex"
    return None

def main(src, out):
    wb = openpyxl.load_workbook(src)
    ct = wb["Control"]
    ct["D4"] = ("Scenario choice: per-axis picker in rows 9-12 (Scanner pattern); "
                "column K inherits the axis choice unless column J holds a per-line override.")
    ct["D4"].font = ARIAL
    ct["G8"] = "Scenario axes: case choice (1-5) per axis"; ct["G8"].font = BOLD
    for axis, r in AXIS_ROW.items():
        ct.cell(row=r, column=7, value=axis).font = ARIAL
        ct.cell(row=r, column=11, value=1).font = BOLD           # K: axis case choice
    ct["G13"] = "Capx axis reserved (Capex block increment)"; ct["G13"].font = ARIAL
    ct["I38"] = "Override"; ct["I38"].font = BOLD
    ct["J38"] = "(blank = inherit axis)"; ct["J38"].font = ARIAL
    n, unmapped = 0, []
    for r in range(14, 400):
        lab = ct.cell(row=r, column=7).value
        if not isinstance(lab, str) or r <= 13 or lab in AXIS_ROW:
            continue
        ax = axis_of(lab)
        if ax is None:
            if ct.cell(row=r, column=11).value is not None:
                unmapped.append((r, lab))
            continue
        ct.cell(row=r, column=11,
                value=f'=IF($J${r}="",K${AXIS_ROW[ax]},$J${r})').font = ARIAL
        n += 1
    wb.properties.creator = wb.properties.lastModifiedBy = "Avia Solutions"
    wb.save(out)
    print(f"axis inheritance applied to {n} lines; unmapped: {unmapped or 'none'}")
    print("saved", out)

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
