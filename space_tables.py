"""Generator v1 increment 10: T2 sqm allocation and T3 busy hour tables (schema
11) with the WS12 tier-2 checks they unlock. Emits a Space & Ops sheet: per-
category commercial sqm series (typed or derived, rule id shown, values carried
forward between typed years), terminal total from T2 reconciled to the Capex
sheet path, busy-hour and ops rows, and the stand-capacity calculation. Appends
tier-2 rules to the existing Checks sheet: R201 stand capacity (physics/ops,
inputs named), R202 busy hour vs annual (working-assumption band until the Peak
Hour data lift), R203 sqm additivity and terminal reconciliation, R204 per-sqm
driver cover (build-time assertion against T1), R208 capex precedes capacity
steps by the lead time. Nothing fabricated: categories without T2 rows stay
blank and R204 reports them. Author: Avia Solutions."""
import sys, re, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ARIAL=Font(name="Arial",size=10); BOLD=Font(name="Arial",size=10,bold=True)
FC=12

def read_header(p):
    h={}
    for line in open(p,encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"): continue
        k,v=line.rstrip("\n").split("\t"); h[k]=v
    return h

def read_tsv(p,cols):
    out=[]
    for line in open(p,encoding="utf-8"):
        if line.startswith("#") or not line.strip(): continue
        out.append(dict(zip(cols,line.rstrip("\n").split("\t"))))
    return out

def main(src,hdr,t2_path,t3_path,t1_path,out):
    H=read_header(hdr); Y0=int(H["start_year"]); LA=int(H["last_actual_year"]); N=int(H["model_term_years"])
    cols=[get_column_letter(FC+i) for i in range(N)]
    t2=read_tsv(t2_path,["space_category","year","case_id","sqm","mode","rule_id","source"])
    t3=read_tsv(t3_path,["year","case_id","metric","value","source"])
    t1=read_tsv(t1_path,["metric_code","segment","case_id","year","value","unit",
                         "temporality","driver_type","step_date","step_value","repeat_years","source"])
    wb=openpyxl.load_workbook(src)
    cx=wb["Capex"]; cxrow={}
    for row in cx.iter_rows(min_col=4,max_col=4):
        if row[0].value=="Terminal size - total [sqm]": cxrow["tt"]=row[0].row
        if row[0].value=="Terminal size - additions [sqm]": cxrow["ad"]=row[0].row
        if row[0].value=="Total expansion (growth) capex": cxrow["gc"]=row[0].row
    if "Space & Ops" in wb.sheetnames: del wb["Space & Ops"]
    so=wb.create_sheet("Space & Ops",wb.sheetnames.index("Capex")+1)
    so["A1"]="=Control!B1"; so["A1"].font=BOLD
    so["B2"]=("Space and operations tables (schema 11 T2 and T3). Typed values carry forward until "
              "the next typed year; derived rows would arrive from the capacity provision engine "
              "(WS6) with their rule id. All demo values working assumptions, illustrative.")
    so["B2"].font=BOLD
    for j in range(N):
        y=Y0+j
        so.cell(row=6,column=FC+j,value=f"{y}{'A' if y<=LA else 'F'}").font=BOLD
    so.cell(row=5,column=46,value="Mode / rule").font=BOLD
    so.cell(row=5,column=48,value="Source").font=BOLD
    # T2: category rows, carried forward between typed years
    cats=sorted({r["space_category"] for r in t2 if r["space_category"]!="terminal_total"})
    r=8; catrow={}
    so.cell(row=r,column=3,value="Commercial space by category [sqm] (T2)").font=BOLD; r+=1
    def series_from(rows_):
        pts={int(x["year"]):(float(x["sqm"]),x["mode"],x["rule_id"],x["source"]) for x in rows_}
        vals=[None]*N; cur=None
        for j in range(N):
            y=Y0+j
            if y in pts: cur=pts[y][0]
            vals[j]=cur
        return vals,pts
    for cat in cats:
        rows_=[x for x in t2 if x["space_category"]==cat]
        vals,pts=series_from(rows_)
        so.cell(row=r,column=4,value=cat).font=ARIAL
        for j in range(N):
            if vals[j] is not None:
                so.cell(row=r,column=FC+j,value=vals[j]).font=ARIAL
        modes="; ".join(f"{y}: {p[1]}"+(f" ({p[2]})" if p[2] not in("-","") else "") for y,p in sorted(pts.items()))
        so.cell(row=r,column=46,value=modes).font=ARIAL
        so.cell(row=r,column=48,value="Source: "+rows_[0]["source"]).font=ARIAL
        catrow[cat]=r; r+=1
    tot=r
    so.cell(row=tot,column=4,value="Total commercial sqm (sum of categories)").font=BOLD
    for j,c in enumerate(cols):
        so.cell(row=tot,column=FC+j,value="="+"+".join(f"{c}{catrow[k]}" for k in cats)).font=ARIAL
    t2term=[x for x in t2 if x["space_category"]=="terminal_total"]
    tvals,tpts=series_from(t2term)
    tterm=tot+1
    so.cell(row=tterm,column=4,value="Terminal total sqm (T2)").font=ARIAL
    for j in range(N):
        if tvals[j] is not None: so.cell(row=tterm,column=FC+j,value=tvals[j]).font=ARIAL
    so.cell(row=tterm,column=48,value="Source: "+ (t2term[0]["source"] if t2term else "none")).font=ARIAL
    trec=tot+2
    so.cell(row=trec,column=4,value="Terminal total per Capex sheet (reconciliation)").font=ARIAL
    for c in cols:
        so.cell(row=trec,column=FC+cols.index(c),value=f"=Capex!{c}{cxrow['tt']}").font=ARIAL
    # T3 rows
    r=trec+2
    so.cell(row=r,column=3,value="Busy hour and operations (T3)").font=BOLD; r+=1
    metrics=["busy_hour_pax","atm_peak_day","stand_count","avg_turnaround_minutes",
             "operating_minutes_per_day","capex_lead_time_years"]
    mrow={}
    for m in metrics:
        rows_=[x for x in t3 if x["metric"]==m]
        if not rows_: continue
        pts={int(x["year"]):float(x["value"]) for x in rows_}
        so.cell(row=r,column=4,value=m).font=ARIAL
        cur=None
        for j in range(N):
            y=Y0+j
            if y in pts: cur=pts[y]
            if cur is not None: so.cell(row=r,column=FC+j,value=cur).font=ARIAL
        so.cell(row=r,column=48,value="Source: "+rows_[0]["source"]).font=ARIAL
        mrow[m]=r; r+=1
    # stand capacity calc row (R201 mechanism, visible)
    sc=r+1
    so.cell(row=sc,column=4,value="Stands required: atm_peak_day x turnaround / operating minutes").font=ARIAL
    for c in cols:
        so.cell(row=sc,column=FC+cols.index(c),
                value=f"=IFERROR({c}{mrow['atm_peak_day']}*{c}{mrow['avg_turnaround_minutes']}/{c}{mrow['operating_minutes_per_day']},0)").font=ARIAL
    # R204 build-time assertion: per_sqm T1 rows need T2 cover
    per_sqm_codes={x["metric_code"] for x in t1 if x["driver_type"]=="per_sqm"}
    uncovered=sorted(per_sqm_codes-set(cats))
    # append tier-2 rules to Checks
    ck=wb["Checks"]
    orow=[x[0].row for x in ck.iter_rows(min_col=2,max_col=2) if x[0].value=="Overall model status"][0]
    c0,cZ=cols[0],cols[-1]
    import datetime
    today=datetime.date.today().strftime("%d %B %Y")
    lead=next((float(x["value"]) for x in t3 if x["metric"]=="capex_lead_time_years"),1.0)
    tier2=[
     ("R201","Stand capacity: stands required within stand count, every populated year",
      f"=IF(SUMPRODUCT(--('Space & Ops'!{c0}{sc}:{cZ}{sc}>'Space & Ops'!{c0}{mrow['stand_count']}:{cZ}{mrow['stand_count']}))=0,\"Ok\",\"ERROR\")",
      "physics/ops (turnaround and operating window named in T3)"),
     ("R202","Busy-hour pax vs annual pax inside 0.03-0.07 percent band, years with pax only (working assumption until the Peak Hour data lift)",
      f"=IF(SUMPRODUCT(--('Space & Ops'!{c0}{mrow['busy_hour_pax']}:{cZ}{mrow['busy_hour_pax']}>Aero!{c0}11:{cZ}11*1000000*0.0007),--(Aero!{c0}11:{cZ}11>0))"
      f"+SUMPRODUCT(--('Space & Ops'!{c0}{mrow['busy_hour_pax']}:{cZ}{mrow['busy_hour_pax']}<Aero!{c0}11:{cZ}11*1000000*0.0003),--(Aero!{c0}11:{cZ}11>0))=0,\"Ok\",\"WATCH\")",
      "assumption (history band pending)"),
     ("R203","Commercial sqm within terminal total, and T2 terminal reconciles to the Capex path",
      f"=IF(AND(SUMPRODUCT(--('Space & Ops'!{c0}{tot}:{cZ}{tot}>'Space & Ops'!{c0}{tterm}:{cZ}{tterm}))=0,"
      f"SUMPRODUCT(--(ABS('Space & Ops'!{c0}{tterm}:{cZ}{tterm}-'Space & Ops'!{c0}{trec}:{cZ}{trec})>0.5))=0),\"Ok\",\"ERROR\")",
      "identity"),
     ("R204",f"Every per-sqm T1 driver row has a T2 allocation ({len(per_sqm_codes)} per-sqm lines checked at build, {today})"
      +(f"; UNCOVERED: {', '.join(uncovered)}" if uncovered else ""),
      f'="{ "Ok" if not uncovered else "ERROR" }"',"generator assertion"),
     ("R208",f"Growth capex present {int(lead)} year(s) (T3 lead time) before each terminal sqm step",
      f"=IF(SUMPRODUCT(--(Capex!{cols[int(lead)]}{cxrow['ad']}:{cZ}{cxrow['ad']}>0),"
      f"--(Capex!{c0}{cxrow['gc']}:{get_column_letter(FC+N-1-int(lead))}{cxrow['gc']}<=0))=0,\"Ok\",\"ERROR\")",
      "physics/ops (lead time named in T3)"),
    ]
    ck.insert_rows(orow,amount=len(tier2)+1)
    ck.cell(row=orow,column=2,value="Tier 2 (T2/T3 tables)").font=BOLD
    for k,(rid,lab,f,basis) in enumerate(tier2):
        rr=orow+1+k
        ck.cell(row=rr,column=2,value=rid).font=BOLD
        ck.cell(row=rr,column=3,value=lab).font=ARIAL
        ck.cell(row=rr,column=7,value=f).font=ARIAL
        ck.cell(row=rr,column=9,value=basis).font=ARIAL
    no=orow+len(tier2)+1
    ck.cell(row=no,column=7,value=f'=IF(COUNTIF(G5:G{no-1},"ERROR")=0,IF(COUNTIF(G5:G{no-1},"WATCH")=0,"Ok","WATCH"),"ERROR")').font=BOLD
    wb.properties.creator=wb.properties.lastModifiedBy="Avia Solutions"
    wb.save(out)
    print(f"saved {out} | T2 categories: {cats} | R204 uncovered: {uncovered or 'none'} | tier-2 rules appended before Checks row {orow}")

if __name__=="__main__":
    main(*sys.argv[1:])
