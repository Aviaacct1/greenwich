"""
Avia standard airport business-plan model: skeleton generator v1.
Extends generate_skeleton_v0 with (a) the year spine read from
project_header.tsv (no calendar year hard-coded in code), and (b) the
granularity tier flag per blueprint 06 section 8 consequence 1: tier 1
pre-DD (Plovdiv line set, one calc sheet per axis), tier 2 DD-grade
(TAS v9 pattern: segment-split aero lines, calc sheets split by category
group, SPV cost line). Line sets come from line_sets.tsv; same grammar,
same taxonomy codes, one flag. Operations Summary keeps rows 8-12 as the
aggregate P&L at both tiers so downstream increments (T1 reader, actuals
strip, financing group, output suite) work unchanged; tier 2 adds group
memo rows below. Author: Avia Solutions.
"""
import sys, openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
FIRST_YCOL = 12
CASES = ["Base", "Spare 1", "Spare 2", "Spare 3", "Spare 4"]

def read_header(path):
    h = {}
    for line in open(path, encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"):
            continue
        k, v = line.rstrip("\n").split("\t")
        h[k] = v
    return h

def read_line_sets(path, tier):
    rows = []
    for line in open(path, encoding="utf-8"):
        if line.startswith("#") or not line.strip():
            continue
        p = line.rstrip("\n").split("\t")
        d = dict(zip(["tier", "axis", "label", "metric_code", "segment", "unit",
                      "driver", "calc_group", "replaces"], p))
        d["tier"] = int(d["tier"])
        rows.append(d)
    if tier == 1:
        return [r for r in rows if r["tier"] == 1]
    replaced = {r["replaces"] for r in rows if r["tier"] == 2 and r["replaces"] != "-"}
    keep = []
    for r in rows:
        if r["tier"] == 1 and f"{r['metric_code']}/{r['segment']}" in replaced:
            continue
        keep.append(r)
    return keep

class G:
    """generation context"""
    def __init__(self, header, lines):
        self.Y0 = int(header["start_year"])
        self.N = int(header["model_term_years"])
        self.jLA = int(header.get("last_actual_year", self.Y0 - 1)) - self.Y0  # actuals-era columns (blueprint 9a)
        self.cols = [get_column_letter(FIRST_YCOL + i) for i in range(self.N)]
        self.lines = lines
        self.sel = {}

def input_block(g, inp, ctrl, label, code, unit, ctrl_row, base_value=0):
    r0 = inp.max_row + 2
    ctrl.cell(row=ctrl_row, column=7, value=label).font = ARIAL
    ctrl.cell(row=ctrl_row, column=11, value=1).font = ARIAL
    inp.cell(row=r0, column=4, value=f"=Control!$G${ctrl_row}").font = BOLD
    for i in range(5):
        r = r0 + 1 + i
        lab = "Base case" if i == 0 else f"=Control!${get_column_letter(13+i-1)}$6"
        inp.cell(row=r, column=6, value=lab).font = ARIAL
        inp.cell(row=r, column=7, value=unit).font = ARIAL
        for j, col in enumerate(g.cols):
            cell = inp[f"{col}{r}"]
            if i == 0:
                cell.value = base_value if j == 0 else f"={g.cols[j-1]}{r}"
            else:
                cell.value = f"={col}{r-1}"
            cell.font = ARIAL
    rs = r0 + 6
    inp.cell(row=rs, column=6, value=f"=F{r0+1}").font = BOLD
    inp.cell(row=rs, column=7, value=unit).font = ARIAL
    inp.cell(row=rs, column=9, value=f"=Control!$K${ctrl_row}").font = ARIAL
    for col in g.cols:
        inp[f"{col}{rs}"] = f"=INDEX({col}{r0+1}:{col}{r0+5},$I{rs})"
        inp[f"{col}{rs}"].font = ARIAL
    g.sel[code] = rs
    return rs

def header_block(g, ws):
    ws["A1"] = "=Control!B1"; ws["A1"].font = BOLD
    ws["A2"] = "=Control!B2"; ws["A2"].font = ARIAL
    for j in range(g.N):
        ws.cell(row=6, column=FIRST_YCOL + j, value=g.Y0 + j).font = BOLD
    traffic = [r for r in g.lines if r["axis"] == "Traffic" and r["metric_code"].startswith("pax_")]
    atms = [r for r in g.lines if r["axis"] == "Traffic" and r["metric_code"].startswith("atm_")]
    ws["D8"] = "Traffic (m pax, two-way)"; ws["D8"].font = BOLD
    for k, r_ in enumerate(traffic):
        r = 9 + k
        ws.cell(row=r, column=4, value=r_["label"]).font = ARIAL
        for col in g.cols:
            ws[f"{col}{r}"] = f"=Inputs!{col}{g.sel[r_['metric_code']]}"; ws[f"{col}{r}"].font = ARIAL
    ws["D11"] = "Total"; ws["D11"].font = BOLD
    for col in g.cols:
        ws[f"{col}11"] = f"=SUM({col}9:{col}10)"; ws[f"{col}11"].font = ARIAL
    ws["D13"] = "ATMs (k, two-way)"; ws["D13"].font = BOLD
    for k, r_ in enumerate(atms):
        r = 14 + k
        ws.cell(row=r, column=4, value=r_["label"]).font = ARIAL
        for col in g.cols:
            ws[f"{col}{r}"] = f"=Inputs!{col}{g.sel[r_['metric_code']]}"; ws[f"{col}{r}"].font = ARIAL
    ws["D18"] = "Total ATMs"; ws["D18"].font = BOLD
    for col in g.cols:
        ws[f"{col}18"] = f"=SUM({col}14:{col}17)"; ws[f"{col}18"].font = ARIAL
    return 20

DRIVER_ROW = {"pax_total": 11, "pax_dom": 9, "pax_intl": 10,
              "atm_dom": 14, "atm_intl": 15, "atm_cargo": 16, "atm_total": 18}

def build_cover(wb, tier):
    ws = wb.create_sheet("Cover")
    ws["C6"] = "Project [Codename]"; ws["C6"].font = Font(name="Arial", size=20, bold=True)
    ws["C8"] = f"Financial Model - skeleton v1, tier {tier} (generated)"; ws["C8"].font = BOLD
    ws["C10"] = "Strictly Private and Confidential"; ws["C10"].font = ARIAL
    ws["C12"] = ("This publication provides general information and should not be relied upon in "
                 "substitution for the exercise of independent judgment. Avia accepts no liability "
                 "of any kind for loss arising from the use of the material presented.")
    ws["C12"].font = ARIAL; ws["C12"].alignment = Alignment(wrap_text=True)
    ws["C14"] = "Copyright Avia Solutions Limited. All rights reserved."; ws["C14"].font = ARIAL
    return ws

def build_control(wb):
    ws = wb.create_sheet("Control")
    ws["B1"] = "[Project]"; ws["B1"].font = BOLD
    for i, cl in enumerate(CASES):
        ws.cell(row=6, column=12 + i, value=cl).font = BOLD
    ws["J6"] = 1
    ws["B2"] = '="Running Case: "&INDEX($L$6:$P$6,$J$6)'
    ws["D4"] = "Case Choice (1-5) is set per line in column K below"; ws["D4"].font = ARIAL
    return ws

def build_aero(g, wb):
    ws = wb.create_sheet("Aero"); r = header_block(g, ws)
    ws.cell(row=r, column=3, value="Aeronautical revenue").font = BOLD; r += 2
    aero_lines = [x for x in g.lines if x["axis"] == "Aero"]
    rev_r, first, last = [], None, None
    for r_ in aero_lines:
        ws.cell(row=r, column=4, value=r_["label"]).font = ARIAL
        ws.cell(row=r, column=7, value="[EUR m]").font = ARIAL
        rev_r.append(r); first = first or r; last = r; r += 1
    total = r + 1
    ws.cell(row=total, column=4, value="Total aeronautical revenue").font = BOLD
    for col in g.cols:
        ws[f"{col}{total}"] = f"=SUM({col}{first}:{col}{last})"; ws[f"{col}{total}"].font = ARIAL
    # regulatory reset factor band, below the total so existing rows never shift.
    # A cumulative product of the reset uplift (one_off_step events on the Inputs
    # reset block): a rate reset in a given year persists thereafter.
    rb = total + 2
    ws.cell(row=rb, column=3, value="Regulatory reset factors (cumulative; step events from Inputs)").font = BOLD
    rb += 1
    reset_r = []
    for r_ in aero_lines:
        upl = g.sel[f"reset:{r_['metric_code']}:{r_['segment']}"]
        ws.cell(row=rb, column=4, value=f"{r_['label']} - reset factor").font = ARIAL
        for j, col in enumerate(g.cols):
            base = "1" if j == 0 else f"{g.cols[j-1]}{rb}"
            ws[f"{col}{rb}"] = f"={base}*(1+Inputs!{col}{upl})"; ws[f"{col}{rb}"].font = ARIAL
        reset_r.append(rb); rb += 1
    # revenue = driver x unit rate x cumulative reset factor
    for r_, rr, rf in zip(aero_lines, rev_r, reset_r):
        sel = g.sel[f"charge:{r_['metric_code']}:{r_['segment']}"]
        for col in g.cols:
            ws[f"{col}{rr}"] = f"={col}{DRIVER_ROW[r_['driver']]}*Inputs!{col}{sel}*{col}{rf}"
            ws[f"{col}{rr}"].font = ARIAL
    return [("Aero", total)]

def build_cat_sheets(g, wb, axis, single_name, block_fn):
    """tier 1: one sheet; tier 2: one sheet per calc_group, TSV order."""
    cats = [x for x in g.lines if x["axis"] == axis]
    groups = []
    for c in cats:
        if c["calc_group"] not in [n for n, _ in groups]:
            groups.append((c["calc_group"], []))
        dict(groups)[c["calc_group"]].append(c)
    if g.tier == 1:
        return [block_fn(g, wb, single_name, cats)]
    return [block_fn(g, wb, name, members) for name, members in groups]

def nonaero_sheet(g, wb, name, cats):
    ws = wb.create_sheet(name); r = header_block(g, ws)
    ws.cell(row=r, column=3, value="Non-aeronautical revenue").font = BOLD; r += 2
    totals = []
    for r_ in cats:
        code = r_["metric_code"]
        ws.cell(row=r, column=3, value=r_["label"]).font = BOLD
        rows = {"paxg": r+1, "epax": r+2, "gpax": r+3, "term": r+4, "eterm": r+5,
                "gterm": r+6, "uplift": r+7, "tot": r+8, "rev": r+9}
        labels = ["Passenger growth", "Elasticity to passenger growth", "Growth from elasticity to traffic",
                  "Terminal size growth", "Elasticity to terminal size growth", "Growth from elasticity terminal size",
                  "Uplift", "Total uplift", "Revenue"]
        for i, l in enumerate(labels):
            ws.cell(row=r+1+i, column=4, value=l).font = ARIAL
        for j, col in enumerate(g.cols):
            pc = g.cols[j-1] if j else None
            ws[f"{col}{rows['paxg']}"] = 0 if j == 0 else f"=IFERROR({col}11/{pc}11-1,0)"
            ws[f"{col}{rows['epax']}"] = f"=Inputs!{col}{g.sel[f'el_pax:{code}']}"
            ws[f"{col}{rows['gpax']}"] = f"={col}{rows['paxg']}*{col}{rows['epax']}"
            ws[f"{col}{rows['term']}"] = 0
            ws[f"{col}{rows['eterm']}"] = f"=Inputs!{col}{g.sel[f'el_term:{code}']}"
            ws[f"{col}{rows['gterm']}"] = f"={col}{rows['term']}*{col}{rows['eterm']}"
            ws[f"{col}{rows['uplift']}"] = f"=Inputs!{col}{g.sel[f'uplift:{code}']}"
            ws[f"{col}{rows['tot']}"] = (f"=(1+{col}{rows['gpax']})*(1+{col}{rows['gterm']})"
                                          f"*(1+{col}{rows['uplift']})-1")
            if j <= max(0, g.jLA):
                # actuals strip: Inputs-fed for actual years, chain only thereafter (blueprint 9a)
                ws[f"{col}{rows['rev']}"] = f"=Inputs!{col}{g.sel[f'baserev:{code}']}"
            else:
                ws[f"{col}{rows['rev']}"] = f"={pc}{rows['rev']}*(1+{col}{rows['tot']})"
            for k in rows.values():
                ws[f"{col}{k}"].font = ARIAL
        totals.append(rows["rev"]); r = rows["rev"] + 2
    ws.cell(row=r, column=4, value="Total non-aeronautical revenue").font = BOLD
    for col in g.cols:
        ws[f"{col}{r}"] = "=" + "+".join(f"{col}{t}" for t in totals); ws[f"{col}{r}"].font = ARIAL
    return (name, r)

def opex_sheet(g, wb, name, cats):
    ws = wb.create_sheet(name); r = header_block(g, ws)
    ws.cell(row=r, column=3, value="Operating costs").font = BOLD; r += 2
    first, last = None, None
    for r_ in cats:
        code = r_["metric_code"]
        ws.cell(row=r, column=4, value=r_["label"]).font = ARIAL
        ws.cell(row=r, column=7, value="[EUR m]").font = ARIAL
        for j, col in enumerate(g.cols):
            if j <= max(0, g.jLA):
                ws[f"{col}{r}"] = f"=Inputs!{col}{g.sel[f'base:{code}']}"   # actuals strip (9a)
            else:
                pc = g.cols[j-1]
                # elasticity to pax growth, then any step event (outsourcing/restructuring);
                # the step chains, so a one-off uplift persists.
                ws[f"{col}{r}"] = (f"={pc}{r}*(1+IFERROR({col}11/{pc}11-1,0)*"
                                   f"Inputs!{col}{g.sel[f'el_pax:{code}']})"
                                   f"*(1+Inputs!{col}{g.sel[f'stepuplift:{code}']})")
            ws[f"{col}{r}"].font = ARIAL
        first = first or r; last = r; r += 1
    ws.cell(row=r + 1, column=4, value="Total operating costs").font = BOLD
    for col in g.cols:
        ws[f"{col}{r+1}"] = f"=SUM({col}{first}:{col}{last})"; ws[f"{col}{r+1}"].font = ARIAL
    return (name, r + 1)

def build_summary_returns(g, wb, aero, nonaero, opex):
    ws = wb.create_sheet("Operations Summary")
    ws["A1"] = "=Control!B1"; ws["A1"].font = BOLD
    for j in range(g.N):
        ws.cell(row=6, column=FIRST_YCOL + j, value=g.Y0 + j).font = BOLD
    def sumref(parts):
        return "=" + "+".join(f"'{n}'!{{c}}{r}" for n, r in parts)
    rows = [("Aeronautical revenue", sumref(aero)),
            ("Non-aeronautical revenue", sumref(nonaero)),
            ("Total revenue", "={c}8+{c}9"),
            ("Operating costs", sumref(opex)),
            ("EBITDA", "={c}10-{c}11")]
    for i, (lbl, f) in enumerate(rows):
        r = 8 + i
        ws.cell(row=r, column=4, value=lbl).font = BOLD if lbl in ("Total revenue", "EBITDA") else ARIAL
        for col in g.cols:
            ws[f"{col}{r}"] = f.format(c=col); ws[f"{col}{r}"].font = ARIAL
    if g.tier == 2:
        memo = 14
        ws.cell(row=memo, column=4, value="Memo: category group totals").font = BOLD
        for k, (n, rr) in enumerate(nonaero + opex):
            r = memo + 1 + k
            ws.cell(row=r, column=4, value=n).font = ARIAL
            for col in g.cols:
                ws[f"{col}{r}"] = f"='{n}'!{col}{rr}"; ws[f"{col}{r}"].font = ARIAL
    rt = wb.create_sheet("Returns")
    rt["A1"] = "=Control!B1"; rt["A1"].font = BOLD
    rt["D8"] = "Returns (v0 stub: entry/exit multiples on EBITDA, per blueprint section 4)"
    rt["D8"].font = ARIAL
    rt["D10"] = "EBITDA entry multiple [x]"; rt["H10"] = 10
    rt["D11"] = "EBITDA exit multiple [x]"; rt["H11"] = 10
    rt["D12"] = "Entry year"; rt["H12"] = g.Y0 + 1
    rt["D13"] = "Exit year"; rt["H13"] = g.Y0 + 10
    for r in range(10, 14):
        rt[f"D{r}"].font = ARIAL; rt[f"H{r}"].font = ARIAL
    return ws

def main(line_sets_path, header_path, tier, out_path):
    header = read_header(header_path)
    lines = read_line_sets(line_sets_path, tier)
    g = G(header, lines); g.tier = tier
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    build_cover(wb, tier); ctrl = build_control(wb)
    inp = wb.create_sheet("Inputs")
    inp["A1"] = "=Control!B1"; inp["A1"].font = BOLD
    for j in range(g.N):
        inp.cell(row=6, column=FIRST_YCOL + j, value=g.Y0 + j).font = BOLD
    ctrl_row = 40
    for r_ in [x for x in lines if x["axis"] == "Traffic"]:
        input_block(g, inp, ctrl, r_["label"], r_["metric_code"], r_["unit"], ctrl_row); ctrl_row += 1
    for r_ in [x for x in lines if x["axis"] == "Aero"]:
        input_block(g, inp, ctrl, f"{r_['label']} - unit rate",
                    f"charge:{r_['metric_code']}:{r_['segment']}", r_["unit"], ctrl_row); ctrl_row += 1
        input_block(g, inp, ctrl, f"{r_['label']} - reset uplift",
                    f"reset:{r_['metric_code']}:{r_['segment']}", "[%]", ctrl_row); ctrl_row += 1
    for r_ in [x for x in lines if x["axis"] == "Non-aero"]:
        code = r_["metric_code"]
        input_block(g, inp, ctrl, f"{r_['label']} - base year revenue", f"baserev:{code}", "[EUR m]", ctrl_row); ctrl_row += 1
        input_block(g, inp, ctrl, f"{r_['label']} - elasticity to pax growth", f"el_pax:{code}", "[x]", ctrl_row); ctrl_row += 1
        input_block(g, inp, ctrl, f"{r_['label']} - elasticity to terminal size", f"el_term:{code}", "[x]", ctrl_row); ctrl_row += 1
        input_block(g, inp, ctrl, f"{r_['label']} - uplift", f"uplift:{code}", "[%]", ctrl_row); ctrl_row += 1
    for r_ in [x for x in lines if x["axis"] == "Opex"]:
        code = r_["metric_code"]
        input_block(g, inp, ctrl, f"{r_['label']} - base year", f"base:{code}", "[EUR m]", ctrl_row); ctrl_row += 1
        input_block(g, inp, ctrl, f"{r_['label']} - elasticity to pax growth", f"el_pax:{code}", "[x]", ctrl_row); ctrl_row += 1
        input_block(g, inp, ctrl, f"{r_['label']} - step uplift", f"stepuplift:{code}", "[%]", ctrl_row); ctrl_row += 1
    aero = build_aero(g, wb)
    nonaero = build_cat_sheets(g, wb, "Non-aero", "Non-aero", nonaero_sheet)
    opex = build_cat_sheets(g, wb, "Opex", "Opex", opex_sheet)
    build_summary_returns(g, wb, aero, nonaero, opex)
    wb.properties.creator = "Avia Solutions"
    wb.properties.lastModifiedBy = "Avia Solutions"
    wb.properties.title = f"Avia Standard Airport Business-Plan Model - skeleton v1 tier {tier}"
    wb.save(out_path)
    print("saved", out_path, "| tier", tier, "| lines", len(lines))

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], int(sys.argv[3]), sys.argv[4])
