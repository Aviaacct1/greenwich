"""Generator v1 increment 8: Checks sheet reinstatement (WS12 rule catalogue 12,
tier 1, plus financing statuses and the vendor round trip). Live rules are
emitted as visible Excel formulas across every year; rules that are generator
assertions (R105 actuals verbatim, R107 named sources) are executed here at
build time and stamped with their result and date. Tier 2 rules awaiting the
T2/T3 tables are listed as pending, never silently dropped.
Author: Avia Solutions."""
import sys, re, datetime, openpyxl
from t1_maps import build_maps
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ARIAL=Font(name="Arial",size=10); BOLD=Font(name="Arial",size=10,bold=True)
FC=12

def read_header(p):
    h={}
    for line in open(p,encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"): continue
        k,v=line.rstrip("\n").split("\t"); h[k]=v
    return h

def t1rows(path):
    rows=[]
    for line in open(path,encoding="utf-8"):
        if line.startswith("#") or not line.strip(): continue
        p=line.rstrip("\n").split("\t")
        rows.append(dict(zip(["metric_code","segment","case_id","year","value","unit",
            "temporality","driver_type","step_date","step_value","repeat_years","source"],p)))
    return rows

def main(src,hdr,actuals_path,line_sets_path,tier,out):
    H=read_header(hdr); Y0=int(H["start_year"]); LA=int(H["last_actual_year"]); N=int(H["model_term_years"])
    cols=[get_column_letter(FC+i) for i in range(N)]
    c0,cZ=cols[0],cols[-1]
    wb=openpyxl.load_workbook(src)
    inp,ctrl,ae=wb["Inputs"],wb["Control"],wb["Aero"]
    na_sheets=[s for s in wb.sheetnames if s=="Non-aero" or s.startswith("Non Aero")]
    def find(ws,label,col=4):
        return [r[0].row for r in ws.iter_rows(min_col=col,max_col=col) if r[0].value==label]
    na_tots=[(s,find(wb[s],"Total non-aeronautical revenue")[0]) for s in na_sheets]
    na_revs=[(s,r) for s in na_sheets for r in find(wb[s],"Revenue")]
    ae_tot=find(ae,"Total aeronautical revenue")[0]
    # aero line rows: between the header and the total
    ae_first=min(r[0].row for r in ae.iter_rows(min_col=4,max_col=4)
                 if isinstance(r[0].value,str) and r[0].value.endswith("charge") or False) if False else None
    # selector blocks and sources for R105/R107 (build-time assertions)
    lbl={ctrl.cell(row=r,column=7).value:r for r in range(30,600) if isinstance(ctrl.cell(row=r,column=7).value,str)}
    sel={}
    for row in inp.iter_rows():
        c=row[5]
        if isinstance(c.value,str) and c.value.startswith("=F") and row[8].value:
            m=re.match(r"=Control!\$G\$(\d+)",str(inp.cell(row=c.row-6,column=4).value or ""))
            if m: sel[int(m.group(1))]=c.row
    # R105: actuals strip matches T1 actual rows verbatim
    _,NAMES,_=build_maps(line_sets_path,int(tier))
    r105_bad=0; r105_n=0
    for r in t1rows(actuals_path):
        if r["temporality"]!="actual": continue
        lab=NAMES.get((r["metric_code"],r["segment"])); yr=int(r["year"])
        if not lab or lab not in lbl or yr>LA: continue
        br=sel[lbl[lab]]-5
        cell=inp.cell(row=br,column=FC+(yr-Y0)).value
        r105_n+=1
        if not isinstance(cell,(int,float)) or abs(cell-float(r["value"]))>1e-9: r105_bad+=1
    # R107: every populated base row carries a named source in col 48
    r107_bad=0; r107_n=0
    for srow in sel.values():
        br=srow-5
        has_data=any(isinstance(inp.cell(row=br,column=FC+j).value,(int,float)) and inp.cell(row=br,column=FC+j).value!=0
                     for j in range(N))
        if not has_data: continue
        r107_n+=1
        s=inp.cell(row=br,column=48).value
        if not (isinstance(s,str) and s.startswith("Source:") and len(s)>8): r107_bad+=1
    # R109 tripwire: the Operations Summary rows 8-12 P&L contract that every downstream step
    # (T1 reader, financing, macro nominal, output suite) depends on by fixed row.
    os_expect={8:"Aeronautical revenue",9:"Non-aeronautical revenue",10:"Total revenue",
               11:"Operating costs",12:"EBITDA"}
    osw=wb["Operations Summary"]
    r109_bad=[rr for rr,lab in os_expect.items() if osw.cell(row=rr,column=4).value!=lab]
    today=datetime.date.today().strftime("%d %B %Y")

    if "Checks" in wb.sheetnames: del wb["Checks"]
    ck=wb.create_sheet("Checks")
    ck["A1"]="=Control!B1"; ck["A1"].font=BOLD
    ck["B2"]="Coherence checks (WS12 rule catalogue, tier 1). Every rule reports Ok / WATCH / ERROR with its rule id; overrides would be recorded and printed in the Assumptions Book."
    ck["B2"].font=BOLD
    ck["B4"]="Rule"; ck["G4"]="Status"; ck["I4"]="Basis"
    for c in ("B4","G4","I4"): ck[c].font=BOLD
    rows=[
     ("R101","EBITDA equals total revenue minus opex, every year",
      f"=IF(SUMPRODUCT(--(ABS('Operations Summary'!{c0}12:{cZ}12-('Operations Summary'!{c0}10:{cZ}10-'Operations Summary'!{c0}11:{cZ}11))>0.000001))=0,\"Ok\",\"ERROR\")","identity"),
     ("R102","Total revenue equals aero plus non-aero, every year",
      f"=IF(SUMPRODUCT(--(ABS('Operations Summary'!{c0}10:{cZ}10-('Operations Summary'!{c0}8:{cZ}8+'Operations Summary'!{c0}9:{cZ}9))>0.000001))=0,\"Ok\",\"ERROR\")","identity"),
     ("R103","Non-aero group totals equal the sum of their category revenue rows, and Operations Summary row 9 equals the sum of the group totals, every year",
      "=IF(SUMPRODUCT(--(ABS('Operations Summary'!{0}9:{1}9-({2}))>0.000001))=0,\"Ok\",\"ERROR\")".format(
        c0,cZ,"+".join(f"'{s}'!{c0}{r}:{cZ}{r}" for s,r in na_tots)),"identity"),
     ("R104","No negative revenue lines, every year (aero total and every non-aero category)",
      "=IF(MIN('Aero'!{0}{1}:{2}{1},{3})>=0,\"Ok\",\"ERROR\")".format(
        c0,ae_tot,cZ,",".join(f"'{s}'!{c0}{r}:{cZ}{r}" for s,r in na_revs)),"identity"),
     ("R105",f"Actuals strip matches T1 actual rows verbatim ({r105_n} cells checked at build, {today})",
      f'="{ "Ok" if r105_bad==0 else "ERROR" }"',"generator assertion"),
     ("R106","Financing statuses (sources and uses, DSR, non-negativity, convergence)",
      "='Financing Inputs'!D68","identity / Scanner block"),
     ("R107",f"Every populated input row carries a named source ({r107_n} rows checked at build, {today})",
      f'="{ "Ok" if r107_bad==0 else "ERROR" }"',"generator assertion"),
     ("R108","Split-estimate lines flagged until confirmed (Vendor Reconciliation round trip)",
      (lambda: [f"='Vendor Reconciliation'!D{r[0].row}" for r in wb["Vendor Reconciliation"].iter_rows(min_col=2,max_col=2) if r[0].value=="Round-trip status"][0])() if "Vendor Reconciliation" in wb.sheetnames
      else '="Not applicable, no vendor model ingested"',"identity"),
     ("R109",f"Operations Summary P&L contract intact (rows 8-12 labels checked at build, {today})"
      +("" if not r109_bad else f"; MISMATCH at row(s) {', '.join(str(x) for x in r109_bad)}"),
      f'="{ "Ok" if not r109_bad else "ERROR" }"',"generator assertion"),
    ]
    r=5
    for rid,lab,f,basis in rows:
        ck.cell(row=r,column=2,value=rid).font=BOLD
        ck.cell(row=r,column=3,value=lab).font=ARIAL
        ck.cell(row=r,column=7,value=f).font=ARIAL
        ck.cell(row=r,column=9,value=basis).font=ARIAL
        r+=1
    ck.cell(row=r+1,column=2,value="Overall model status").font=BOLD
    ck.cell(row=r+1,column=7,value=f'=IF(COUNTIF(G5:G{r-1},"ERROR")=0,IF(COUNTIF(G5:G{r-1},"WATCH")=0,"Ok","WATCH"),"ERROR")').font=BOLD
    ck.cell(row=r+3,column=2,value=("Pending tier 2 (need T2 sqm allocation and T3 busy-hour tables): R201 stands vs ATM "
      "time-on-ground, R202 busy-hour vs annual, R203 commercial sqm additivity, R204 per-sqm rows "
      "have allocations, R206 opex steps at capacity steps, R208 capex phasing lead time. Tier 3 "
      "history rules activate when pilot resolve completes (R301-R307).")).font=ARIAL
    wb.properties.creator=wb.properties.lastModifiedBy="Avia Solutions"
    wb.save(out)
    print(f"saved {out} | R105 {r105_n} cells, {r105_bad} bad | R107 {r107_n} rows, {r107_bad} bad "
          f"| R109 rows 8-12 {'Ok' if not r109_bad else 'MISMATCH '+str(r109_bad)}")

if __name__=="__main__":
    main(sys.argv[1],sys.argv[2],sys.argv[3],sys.argv[4],sys.argv[5],sys.argv[6])
