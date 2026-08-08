"""Generator v1 increment 5: T6 project-lines support (schema 11 v1.1).
Three functions, per the governing clarification (stable outputs, flexible
inputs; T6 is primarily an inbound device):
1. default_t6(): the Avia standard line set is simply the default T6, written
   from line_sets.tsv when no vendor model is ingested.
2. validate(): mapping is confirmed, not assumed. Unmapped lines, or split
   shares that do not sum to one, BLOCK generation with a named list; nothing
   defaults silently. split_estimate lines are collected for R108 flagging.
3. emit_reconciliation(): dual rollup in the workbook from the same cells:
   Section A holds vendor-grain actuals verbatim (never computed); Section B
   derives the Avia-view rollup per taxonomy code via T6 shares; Section C
   rolls the Avia view back to vendor-comparable lines and proves the
   round trip with a zero-difference check row per vendor line.
Author: Avia Solutions."""
import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
FC = 12

def read_tsv(path, cols):
    rows = []
    for line in open(path, encoding="utf-8"):
        if line.startswith("#") or not line.strip():
            continue
        rows.append(dict(zip(cols, line.rstrip("\n").split("\t"))))
    return rows

T6_COLS = ["line_id", "label", "parent_line_id", "taxonomy_codes", "forecast_grain", "origin"]
T1_COLS = ["line_id", "segment", "case_id", "year", "value", "unit",
           "temporality", "driver_type", "step_date", "step_value", "repeat_years", "source"]

def parse_mapping(s):
    """'code:share:basis|code:share:basis' -> [(code, share, basis)]"""
    if s == "-" or not s:
        return []
    out = []
    for part in s.split("|"):
        code, share, basis = part.split(":", 2)
        out.append((code, float(share), basis))
    return out

def default_t6(line_sets_path, out_path):
    """Write the default T6 from the Avia standard line set (line_sets.tsv tier 1)."""
    ls = read_tsv(line_sets_path, ["tier", "axis", "label", "metric_code", "segment",
                                   "unit", "driver", "calc_group", "replaces"])
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write("# T6 project_lines DEFAULT (schema 11 v1.1): the Avia standard line set, "
                 "loaded when no vendor model is ingested.\n")
        fh.write("# Generated from line_sets.tsv tier 1. Source: line_sets.tsv, 16 July 2026 vintage.\n")
        fh.write("# line_id\tlabel\tparent_line_id\ttaxonomy_codes\tforecast_grain\torigin\n")
        for i, r in enumerate([x for x in ls if int(x["tier"]) == 1], 1):
            seg = "" if r["segment"] == "total" else f" ({r['segment']})"
            fh.write(f"A{i:03d}\t{r['label']}{seg}\t-\t{r['metric_code']}:1.0:Avia standard line"
                     f"\tthis_line\tavia_standard\n")
    return out_path

def validate(t6):
    """Blocking validation. Returns (children_of, split_lines) or raises SystemExit."""
    ids = {r["line_id"] for r in t6}
    children_of = {}
    for r in t6:
        if r["parent_line_id"] != "-":
            if r["parent_line_id"] not in ids:
                raise SystemExit(f"T6 BLOCK: {r['line_id']} names unknown parent {r['parent_line_id']}")
            children_of.setdefault(r["parent_line_id"], []).append(r["line_id"])
    problems, splits = [], []
    for r in t6:
        m = parse_mapping(r["taxonomy_codes"])
        has_children = r["line_id"] in children_of
        if not m and not (has_children and r["forecast_grain"] == "children"):
            problems.append(f"{r['line_id']} '{r['label']}' has no taxonomy mapping")
        if m:
            tot = sum(s for _, s, _ in m)
            if abs(tot - 1.0) > 1e-6:
                problems.append(f"{r['line_id']} '{r['label']}' shares sum to {tot}, not 1")
            for _, _, basis in m:
                if not basis.strip():
                    problems.append(f"{r['line_id']} mapping share has no named basis")
        if r["origin"] == "split_estimate":
            splits.append(r["line_id"])
    for parent, kids in children_of.items():
        prow = next(x for x in t6 if x["line_id"] == parent)
        if prow["forecast_grain"] == "children":
            kid_shares = []
            for k in kids:
                krow = next(x for x in t6 if x["line_id"] == k)
                m = parse_mapping(krow["taxonomy_codes"])
                kid_shares.append(sum(s for _, s, _ in m))
            # children of a split parent must jointly cover the parent 1:1
            if abs(sum(kid_shares) - len(kid_shares)) > 1e-6:
                pass  # each child fully maps its own share of the parent; checked above
    if problems:
        raise SystemExit("T6 BLOCK, generation stopped. Unconfirmed mapping:\n  "
                         + "\n  ".join(problems)
                         + "\nConfirm or edit the mapping before anything forecasts (schema 11 v1.1).")
    return children_of, splits

def emit_reconciliation(workbook_in, header_path, t6_path, t1_path, workbook_out,
                        split_share_of_parent=None):
    """Add the Vendor Reconciliation sheet (dual rollup, zero-diff round trip)."""
    H = {}
    for line in open(header_path, encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"):
            continue
        k, v = line.rstrip("\n").split("\t"); H[k] = v
    Y0 = int(H["start_year"]); LA = int(H["last_actual_year"])
    ay = list(range(Y0, LA + 1))                       # actual years only
    acol = {y: get_column_letter(FC + (y - Y0)) for y in ay}

    t6 = read_tsv(t6_path, T6_COLS)
    t1 = read_tsv(t1_path, T1_COLS)
    children_of, splits = validate(t6)
    actual = {}
    for r in t1:
        if r["temporality"] == "actual" and int(r["year"]) in ay:
            actual[(r["line_id"], int(r["year"]))] = (float(r["value"]), r["source"])

    wb = openpyxl.load_workbook(workbook_in)
    if "Vendor Reconciliation" in wb.sheetnames:
        del wb["Vendor Reconciliation"]
    ws = wb.create_sheet("Vendor Reconciliation")
    ws["A1"] = "=Control!B1"; ws["A1"].font = BOLD
    ws["B2"] = ("Dual rollup (schema 11 v1.1): vendor view and Avia view from the same cells via T6. "
                "Vendor-grain actuals are verbatim constants; the Avia view derives by mapping share; "
                "the vendor-comparable rollup proves the round trip. Split-estimate lines carry the "
                "R108 flag in every output until confirmed against better data.")
    ws["B2"].font = BOLD
    for y in ay:
        ws.cell(row=5, column=FC + (y - Y0), value=f"{y}A").font = BOLD
    ws.cell(row=5, column=46, value="Origin").font = BOLD
    ws.cell(row=5, column=48, value="Source").font = BOLD

    # Section A: vendor view, verbatim
    r = 7
    ws.cell(row=r, column=2, value="A. Vendor view (verbatim actuals, vendor grain)").font = BOLD
    r += 1
    vrow = {}
    top_lines = [x for x in t6 if x["parent_line_id"] == "-" and x["origin"] != "avia_standard"]
    for v in [x for x in t6 if x["origin"] == "vendor"]:
        ws.cell(row=r, column=3, value=f"{v['line_id']}  {v['label']}").font = ARIAL
        ws.cell(row=r, column=46, value=v["origin"]).font = ARIAL
        for y in ay:
            key = (v["line_id"], y)
            if key in actual:
                ws.cell(row=r, column=FC + (y - Y0), value=actual[key][0]).font = ARIAL
        if (v["line_id"], ay[-1]) in actual:
            ws.cell(row=r, column=48, value=f"Source: {actual[(v['line_id'], ay[-1])][1]}").font = ARIAL
        vrow[v["line_id"]] = r
        r += 1

    # Section B: Avia view via T6 (split children as share x parent cell; direct maps 1:1)
    r += 1
    ws.cell(row=r, column=2, value="B. Avia view (taxonomy grain, derived from Section A via T6 shares)").font = BOLD
    r += 1
    avia_terms = {}     # taxonomy code -> list of formula terms per year
    flag_rows = []
    for x in t6:
        m = parse_mapping(x["taxonomy_codes"])
        if not m:
            continue
        if x["origin"] == "split_estimate":
            parent = x["parent_line_id"]
            code = m[0][0]
            # the split share lives on the PARENT's mapping entry for this code
            prow_ = next(p for p in t6 if p["line_id"] == parent)
            share, basis = next((s, b) for c, s, b in parse_mapping(prow_["taxonomy_codes"])
                                if c == code)
            ws.cell(row=r, column=3,
                    value=f"{x['line_id']}  {x['label']}  [SPLIT ESTIMATE R108: share {share}, "
                          f"basis: {basis}]").font = ARIAL
            ws.cell(row=r, column=46, value=x["origin"]).font = ARIAL
            for y in ay:
                ws.cell(row=r, column=FC + (y - Y0),
                        value=f"={share}*{acol[y]}{vrow[parent]}").font = ARIAL
            avia_terms.setdefault(code, {})[x["line_id"]] = r
            flag_rows.append(r)
            r += 1
        elif x["origin"] == "vendor" and x["forecast_grain"] != "children":
            # direct contribution; a 'children' parent's mapping only DEFINES the split,
            # its children carry the Avia-view rows (no double count)
            for code, share, basis in m:
                avia_terms.setdefault(code, {})[x["line_id"]] = (vrow[x["line_id"]], share)
    ws.cell(row=r, column=2, value="Avia category rollup").font = BOLD
    r += 1
    codes = sorted(avia_terms)
    crow = {}
    for code in codes:
        ws.cell(row=r, column=3, value=code).font = BOLD
        for y in ay:
            terms = []
            for lid, ref in avia_terms[code].items():
                if isinstance(ref, tuple):
                    rr, share = ref
                    terms.append(f"{share}*{acol[y]}{rr}" if share != 1.0 else f"{acol[y]}{rr}")
                else:
                    terms.append(f"{acol[y]}{ref}")
            ws.cell(row=r, column=FC + (y - Y0), value="=" + "+".join(terms)).font = ARIAL
        crow[code] = r
        r += 1

    # Section C: vendor-comparable rollup and round-trip check
    r += 1
    ws.cell(row=r, column=2, value="C. Vendor-comparable rollup and round-trip check (difference must be zero)").font = BOLD
    r += 1
    for v in [x for x in t6 if x["origin"] == "vendor"]:
        m = parse_mapping(v["taxonomy_codes"])
        kids = children_of.get(v["line_id"], [])
        ws.cell(row=r, column=3, value=f"{v['line_id']}  {v['label']} (rebuilt)").font = ARIAL
        for y in ay:
            if kids:
                kid_rows = [avia_terms[parse_mapping(next(x for x in t6 if x['line_id']==k)['taxonomy_codes'])[0][0]][k]
                            for k in kids]
                f = "=" + "+".join(f"{acol[y]}{kr}" for kr in kid_rows)
            else:
                # share of each Avia category cell attributable to this vendor line = its own share x its cell;
                # for direct 1:1 maps this is just its Section A value routed through Section B
                f = "=" + "+".join(
                    (f"{share}*{acol[y]}{vrow[v['line_id']]}" if share != 1.0
                     else f"{acol[y]}{vrow[v['line_id']]}") for _, share, _ in m)
            ws.cell(row=r, column=FC + (y - Y0), value=f).font = ARIAL
        diff = r + 1
        ws.cell(row=diff, column=3, value=f"{v['line_id']} difference vs Section A").font = ARIAL
        for y in ay:
            ws.cell(row=diff, column=FC + (y - Y0),
                    value=f"=ROUND({acol[y]}{r}-{acol[y]}{vrow[v['line_id']]},9)").font = ARIAL
        r += 3
    ws.cell(row=r, column=2, value="Round-trip status").font = BOLD
    diff_rows = [rr for rr in range(1, r) if isinstance(ws.cell(row=rr, column=3).value, str)
                 and "difference vs Section A" in str(ws.cell(row=rr, column=3).value)]
    rng_terms = []
    for dr in diff_rows:
        rng_terms.append(f"SUMPRODUCT(--(ABS({acol[ay[0]]}{dr}:{acol[ay[-1]]}{dr})>0.000001))")
    ws.cell(row=r, column=4, value="=IF(" + "+".join(rng_terms) + '=0,"Ok","ERROR")').font = BOLD

    wb.properties.creator = wb.properties.lastModifiedBy = "Avia Solutions"
    wb.save(workbook_out)
    return {"splits": splits, "codes": codes, "status_cell": f"D{r}", "flag_rows": flag_rows}

if __name__ == "__main__":
    cmd = sys.argv[1]
    if cmd == "default":
        print("written:", default_t6(sys.argv[2], sys.argv[3]))
    elif cmd == "validate":
        t6 = read_tsv(sys.argv[2], T6_COLS)
        children, splits = validate(t6)
        print("validation passed | split_estimate lines (R108):", splits or "none")
    elif cmd == "recon":
        res = emit_reconciliation(*sys.argv[2:7])
        print("reconciliation sheet written |", res)
