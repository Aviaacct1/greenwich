"""Actuals strip v2 (generator v1 increment 2, parameterised paths). Rebuilds
A/F year headers on every year-spined sheet, writes T1 actual rows as constants
in the actual-year columns, and stamps sources. Logic unchanged from increment 2;
paths now come from argv and the spine from project_header.tsv.
Author: Avia Solutions."""
import sys, re, openpyxl
from t1_maps import build_maps
from openpyxl.styles import Font
ARIAL=Font(name="Arial",size=10); BOLD=Font(name="Arial",size=10,bold=True)
FC=12

def header(path):
    h={}
    for line in open(path,encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"): continue
        k,v=line.rstrip("\n").split("\t"); h[k]=v
    return h

def t1(path):
    rows=[]
    for line in open(path,encoding="utf-8"):
        if line.startswith("#") or not line.strip(): continue
        p=line.rstrip("\n").split("\t")
        rows.append(dict(zip(["metric_code","segment","case_id","year","value","unit",
            "temporality","driver_type","step_date","step_value","repeat_years","source"],p)))
    return rows

def main(src,hdr,actuals,line_sets_path,tier,out):
    H=header(hdr)
    Y0=int(H["start_year"]); LA=int(H["last_actual_year"]); N=int(H["model_term_years"])
    wb=openpyxl.load_workbook(src)
    inp=wb["Inputs"]
    for ws in wb.worksheets:
        if ws.title in ("Cover","Control","Returns","Valuation","Checks","Charts"): continue
        if ws.cell(row=6,column=FC).value is None: continue
        for j in range(N):
            y=Y0+j
            ws.cell(row=6,column=FC+j,value=f"{y}{'A' if y<=LA else 'F'}").font=BOLD
    ctrl=wb["Control"]
    lbl_row={ctrl.cell(row=r,column=7).value:r for r in range(30,400) if isinstance(ctrl.cell(row=r,column=7).value,str)}
    sel={}
    for row in inp.iter_rows():
        c=row[5]
        if isinstance(c.value,str) and c.value.startswith("=F") and row[8].value:
            m=re.match(r"=Control!\$G\$(\d+)",str(inp.cell(row=c.row-6,column=4).value or ""))
            if m: sel[int(m.group(1))]=c.row
    _,NAMES,_=build_maps(line_sets_path,int(tier))
    n=0
    for r in t1(actuals):
        if r["temporality"]!="actual": continue
        lab=NAMES.get((r["metric_code"],r["segment"])); yr=int(r["year"])
        if not lab or lab not in lbl_row or yr>LA: continue
        br=sel[lbl_row[lab]]-5
        inp.cell(row=br,column=FC+(yr-Y0),value=float(r["value"]))
        inp.cell(row=br,column=48,value=f"Source: {r['source']} (actuals to {LA})")
        n+=1
    wb["Cover"]["C10"]=f"Actuals to {LA} (constants); forecast from {LA+1}; term {N} years from {Y0}. No hard-coded years."
    wb["Cover"]["C10"].font=ARIAL
    wb.properties.creator=wb.properties.lastModifiedBy="Avia Solutions"
    wb.save(out)
    print("actual cells written:",n,"| first forecast year:",LA+1)

if __name__=="__main__":
    main(sys.argv[1],sys.argv[2],sys.argv[3],sys.argv[4],sys.argv[5],sys.argv[6])
