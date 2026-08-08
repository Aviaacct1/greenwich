"""Generator v1 increment 13: the macro feed (John, 16 July: analysts assemble
GDP forecasts and long-term inflation indices per project by hand; the sources
all sit in the Avia estate). Takes one or more macro source files in the common
vendor layout (country rows x year columns: IMF WEO, Oxford Economics, Global
Insight exports all reduce to this), a country, and a metric per file (gdp_growth,
cpi, rpi, construction_index), and emits T1 macro rows: one row set PER SOURCE
with the source and vintage named, plus, where two or more GDP sources are
given, a BASKET row set (weighted mean, weights and basis stated on every row).
Nothing is fabricated: years missing from a source are absent from its rows and
the basket uses only the sources present in that year (membership noted).
Output feeds the workbook Macro block (CPI inflation and cumulative index; the
block's restoration in the header grammar is the queued companion increment)
and the Capex real-to-nominal step (construction index). Author: Avia Solutions."""
import sys, os, re, csv, datetime

YEAR_RE = re.compile(r"^(19[9]\d|20[0-6]\d)$")

def read_grid(path):
    if path.lower().endswith((".csv", ".tsv")):
        d = "\t" if path.lower().endswith(".tsv") else ","
        with open(path, encoding="utf-8-sig") as fh:
            return [r for r in csv.reader(fh, delimiter=d)]
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [[c.value for c in row] for row in ws.iter_rows()]

def series_for(path, country):
    """Find the year header row, then the country row; returns {year: value}."""
    rows = read_grid(path)
    hdr, years = None, {}
    for i, row in enumerate(rows[:30]):
        ys = {}
        for j, v in enumerate(row):
            s = str(v).strip() if v is not None else ""
            if YEAR_RE.match(s) or (isinstance(v, (int, float)) and 1990 <= int(v or 0) <= 2069 and v == int(v)):
                ys[j] = int(float(s))
        if len(ys) >= 3:
            hdr, years = i, ys
            break
    if hdr is None:
        raise SystemExit(f"MACRO BLOCK: no year header found in {path}")
    cl = country.lower()
    for row in rows[hdr + 1:]:
        label = next((str(v).strip() for v in row if isinstance(v, str) and v.strip()), "")
        if label.lower() == cl or cl in label.lower():
            out = {}
            for j, y in years.items():
                v = row[j] if j < len(row) else None
                if isinstance(v, (int, float)):
                    out[y] = float(v)
            if out:
                return label, out
    raise SystemExit(f"MACRO BLOCK: country '{country}' not found in {path}")

def main(country, out_path, *specs):
    """specs: metric=source_name=vintage=path[=weight]  e.g.
    gdp_growth=IMF WEO=Apr 2026=weo.xlsx=0.5"""
    rows, gdp_sources = [], []
    today = datetime.date.today().strftime("%d %B %Y")
    for spec in specs:
        parts = spec.split("=")
        metric, name, vintage, path = parts[0], parts[1], parts[2], parts[3]
        weight = float(parts[4]) if len(parts) > 4 else None
        label, ser = series_for(path, country)
        src = f"{name}, {vintage}, '{label}', {os.path.basename(path)}"
        for y, v in sorted(ser.items()):
            rows.append(f"{metric}\t{name.replace(' ','_').lower()}\t1\t{y}\t{v}\t"
                        f"{'growth' if metric=='gdp_growth' else 'index_pct'}\tforecast\tlevel\t\t\t\t{src}")
        if metric == "gdp_growth":
            gdp_sources.append((name, vintage, ser, weight))
    if len(gdp_sources) >= 2:
        given = [w for _, _, _, w in gdp_sources if w is not None]
        equal = len(given) != len(gdp_sources)
        n = len(gdp_sources)
        all_years = sorted({y for _, _, s, _ in gdp_sources for y in s})
        for y in all_years:
            members = [(nm, s[y], (1.0 / n if equal else w)) for nm, _, s, w in gdp_sources if y in s]
            tw = sum(w for _, _, w in members)
            basket = sum(v * w for _, v, w in members) / tw
            basis = ("equal weights" if equal else "stated weights") + \
                    ", members this year: " + ", ".join(f"{nm} ({round(w/tw,2)})" for nm, _, w in members)
            rows.append(f"gdp_growth\tbasket\t1\t{y}\t{round(basket,4)}\tgrowth\tforecast\tlevel\t\t\t\t"
                        f"Avia basket of {len(members)} source(s): {basis}; assembled {today}")
    hdr = (f"# T1 macro rows for {country} | assembled {today} by macro_feed.py\n"
           "# Segment carries the source; the 'basket' segment is the Avia weighted mean with "
           "weights and membership named per row. Point the specs at the live estate files "
           "(E:\\Avia\\Data, Egnyte 02 Knowledge) when mounted; demo runs use illustrative copies.\n"
           "# metric_code\tsegment\tcase_id\tyear\tvalue\tunit\ttemporality\tdriver_type"
           "\tstep_date\tstep_value\trepeat_years\tsource\n")
    open(out_path, "w", encoding="utf-8").write(hdr + "\n".join(rows) + "\n")
    print(f"macro rows written: {len(rows)} | gdp sources: {len(gdp_sources)}"
          f"{' | basket emitted' if len(gdp_sources)>=2 else ''} | {out_path}")

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], *sys.argv[3:])
