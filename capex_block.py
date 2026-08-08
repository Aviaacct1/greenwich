"""Generator v1 increment 7a: Capex block (blueprint 06 section 4 Capex grammar).
Emits Inputs/Control case blocks for the five expansion capex lines, maintenance
capex and committed terminal sqm additions (all inheriting the Capex axis on the
per-axis picker), builds the Capex calc sheet (growth capex total, maintenance,
terminal sqm path and growth, depreciation from useful lives: existing base over
20 years, new spend straight-line over 25 years with an explicit visible window
per column), and wires the non-aero category blocks' Terminal size growth rows
(zero placeholders since v0) to the Capex terminal growth row. Year spine from
project_header.tsv; no calendar year hard-coded in code; demo spend profile
labelled working assumption. Author: Avia Solutions."""
import sys, re, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ARIAL=Font(name="Arial",size=10); BOLD=Font(name="Arial",size=10,bold=True)
FC=12

def read_header(p):
    h={}
    for line in open(p,encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"): continue
        k,v=line.rstrip("\n").split("\t"); h[k]=v
    return h

CAPEX_LINES=[("Capex - land and buildings","capex_land"),
 ("Capex - machinery, plant and equipment","capex_mpe"),
 ("Capex - furniture and vehicles","capex_fv"),
 ("Capex - other tangibles","capex_other"),
 ("Capex - interim facility","capex_interim")]
# demo spend profile, EUR m, working assumption illustrative (year offsets from entry year)
DEMO={"capex_land":{3:10.0,4:25.0},"capex_mpe":{4:8.0},"capex_fv":{4:1.5},
 "capex_other":{},"capex_interim":{2:2.0}}
SQM_ADD_OFFSET,SQM_ADD=4,6000        # +6,000 sqm four years after entry
OPENING_SQM=20000.0
EXIST_BASE,LIFE_EXIST,LIFE_NEW=150.0,20,25

def input_block(inp,ctrl,cols,label,unit,ctrl_row,axis_row):
    r0=inp.max_row+2
    ctrl.cell(row=ctrl_row,column=7,value=label).font=ARIAL
    ctrl.cell(row=ctrl_row,column=11,value=f'=IF($J${ctrl_row}="",K${axis_row},$J${ctrl_row})').font=ARIAL
    inp.cell(row=r0,column=4,value=f"=Control!$G${ctrl_row}").font=BOLD
    for i in range(5):
        r=r0+1+i
        lab="Base case" if i==0 else f"=Control!${get_column_letter(13+i-1)}$6"
        inp.cell(row=r,column=6,value=lab).font=ARIAL
        inp.cell(row=r,column=7,value=unit).font=ARIAL
        for j,col in enumerate(cols):
            c=inp[f"{col}{r}"]
            c.value=0 if i==0 else f"={col}{r-1}"
            c.font=ARIAL
    rs=r0+6
    inp.cell(row=rs,column=6,value=f"=F{r0+1}").font=BOLD
    inp.cell(row=rs,column=7,value=unit).font=ARIAL
    inp.cell(row=rs,column=9,value=f"=Control!$K${ctrl_row}").font=ARIAL
    for col in cols:
        inp[f"{col}{rs}"]=f"=INDEX({col}{r0+1}:{col}{r0+5},$I{rs})"; inp[f"{col}{rs}"].font=ARIAL
    return r0+1,rs   # base row, selector row

def main(src,hdr,out):
    H=read_header(hdr); Y0=int(H["start_year"]); LA=int(H["last_actual_year"]); N=int(H["model_term_years"])
    ENTRY=LA+1; jE=ENTRY-Y0
    cols=[get_column_letter(FC+i) for i in range(N)]
    wb=openpyxl.load_workbook(src)
    inp,ctrl=wb["Inputs"],wb["Control"]
    # Capex axis on the picker (row 13 was reserved)
    ctrl["G13"]="Capex"; ctrl["G13"].font=ARIAL
    ctrl.cell(row=13,column=11,value=1).font=BOLD
    # next free Control label row
    used=[r for r in range(14,600) if isinstance(ctrl.cell(row=r,column=7).value,str)]
    cr=max(used)+2
    sel={}
    for lab,code in CAPEX_LINES:
        br,rs=input_block(inp,ctrl,cols,lab,"[EUR m]",cr,13); sel[code]=(br,rs); cr+=1
    br,rs=input_block(inp,ctrl,cols,"Maintenance capex","[EUR m]",cr,13); sel["capex_maint"]=(br,rs); cr+=1
    br,rs=input_block(inp,ctrl,cols,"Terminal size - committed additions","[sqm]",cr,13); sel["sqm_add"]=(br,rs); cr+=1
    # demo values into base rows (constants per year, source labelled)
    WA="working assumption, illustrative"
    for code,prof in DEMO.items():
        b=sel[code][0]
        for off,v in prof.items():
            j=jE+off
            if j<N: inp.cell(row=b,column=FC+j,value=v)
        inp.cell(row=b,column=48,value=f"Source: {WA}").font=ARIAL
    b=sel["capex_maint"][0]
    inp.cell(row=b,column=FC+jE,value=3.0)
    for j in range(jE+1,N):
        inp.cell(row=b,column=FC+j,value=f"={cols[j-1]}{b}*1.02")
    for j in range(0,jE): inp.cell(row=b,column=FC+j,value=0)
    inp.cell(row=b,column=48,value=f"Source: {WA}, 2% real drift").font=ARIAL
    b=sel["sqm_add"][0]
    j=jE+SQM_ADD_OFFSET
    if j<N: inp.cell(row=b,column=FC+j,value=SQM_ADD)
    inp.cell(row=b,column=48,value=f"Source: {WA}, committed extension").font=ARIAL

    # Capex calc sheet
    if "Capex" in wb.sheetnames: del wb["Capex"]
    cx=wb.create_sheet("Capex",wb.sheetnames.index("Operations Summary"))
    cx["A1"]="=Control!B1"; cx["A1"].font=BOLD
    cx["B2"]=("Capex, terminal size and depreciation [EUR m; sqm]. Five expansion lines per the "
              "blueprint Capex grammar; depreciation straight-line from useful lives; terminal "
              "size feeds the non-aero elasticity blocks. Demo profile is a working assumption.")
    cx["B2"].font=BOLD
    for j in range(N):
        y=Y0+j
        cx.cell(row=6,column=FC+j,value=f"{y}{'A' if y<=LA else 'F'}").font=BOLD
    rows={}
    r=8
    for lab,code in CAPEX_LINES:
        cx.cell(row=r,column=4,value=lab).font=ARIAL
        for col in cols: cx[f"{col}{r}"]=f"=Inputs!{col}{sel[code][1]}"; cx[f"{col}{r}"].font=ARIAL
        rows[code]=r; r+=1
    gc=r; cx.cell(row=gc,column=4,value="Total expansion (growth) capex").font=BOLD
    for col in cols: cx[f"{col}{gc}"]=f"=SUM({col}8:{col}{r-1})"; cx[f"{col}{gc}"].font=ARIAL
    mt=gc+2; cx.cell(row=mt,column=4,value="Maintenance capex").font=ARIAL
    for col in cols: cx[f"{col}{mt}"]=f"=Inputs!{col}{sel['capex_maint'][1]}"; cx[f"{col}{mt}"].font=ARIAL
    sq0=mt+2
    cx.cell(row=sq0,column=4,value="Terminal size - opening [sqm]").font=ARIAL
    cx.cell(row=sq0,column=8,value=OPENING_SQM).font=BOLD
    cx.cell(row=sq0,column=48,value="Source: working assumption, illustrative").font=ARIAL
    ad=sq0+1; cx.cell(row=ad,column=4,value="Terminal size - additions [sqm]").font=ARIAL
    for col in cols: cx[f"{col}{ad}"]=f"=Inputs!{col}{sel['sqm_add'][1]}"; cx[f"{col}{ad}"].font=ARIAL
    tt=ad+1; cx.cell(row=tt,column=4,value="Terminal size - total [sqm]").font=BOLD
    for j,col in enumerate(cols):
        cx[f"{col}{tt}"]=f"=$H${sq0}+{col}{ad}" if j==0 else f"={cols[j-1]}{tt}+{col}{ad}"
        cx[f"{col}{tt}"].font=ARIAL
    tg=tt+1; cx.cell(row=tg,column=4,value="Terminal size growth [%]").font=ARIAL
    for j,col in enumerate(cols):
        cx[f"{col}{tg}"]=0 if j==0 else f"=IFERROR({col}{tt}/{cols[j-1]}{tt}-1,0)"
        cx[f"{col}{tg}"].font=ARIAL
    ul=tg+2
    cx.cell(row=ul,column=4,value="Useful life - existing assets [years]").font=ARIAL; cx.cell(row=ul,column=8,value=LIFE_EXIST).font=BOLD
    cx.cell(row=ul+1,column=4,value="Useful life - new capex [years]").font=ARIAL; cx.cell(row=ul+1,column=8,value=LIFE_NEW).font=BOLD
    cx.cell(row=ul+2,column=4,value="Existing asset base [EUR m]").font=ARIAL; cx.cell(row=ul+2,column=8,value=EXIST_BASE).font=BOLD
    cx.cell(row=ul+2,column=48,value="Source: working assumption, illustrative (Plovdiv observed lives 20y existing, 25y new)").font=ARIAL
    de=ul+4; cx.cell(row=de,column=4,value="Depreciation - existing assets").font=ARIAL
    for j,col in enumerate(cols):
        cx[f"{col}{de}"]=f"=-IF({j}<$H${ul},$H${ul+2}/$H${ul},0)"; cx[f"{col}{de}"].font=ARIAL
    dn=de+1; cx.cell(row=dn,column=4,value="Depreciation - new capex (straight line, explicit window)").font=ARIAL
    for j,col in enumerate(cols):
        w0=max(0,j-(LIFE_NEW-1))
        cx[f"{col}{dn}"]=f"=-SUM({cols[w0]}{gc}:{col}{gc})/$H${ul+1}"; cx[f"{col}{dn}"].font=ARIAL
    dt=dn+1; cx.cell(row=dt,column=4,value="Total depreciation").font=BOLD
    for col in cols: cx[f"{col}{dt}"]=f"={col}{de}+{col}{dn}"; cx[f"{col}{dt}"].font=ARIAL

    # wire non-aero Terminal size growth rows (constant 0 since v0) to the Capex growth row
    wired=0
    for sn in [s for s in wb.sheetnames if s=="Non-aero" or s.startswith("Non Aero")]:
        na=wb[sn]
        for row in na.iter_rows(min_col=4,max_col=4):
            c=row[0]
            if c.value=="Terminal size growth":
                for col in cols:
                    na[f"{col}{c.row}"]=f"=Capex!{col}{tg}"; na[f"{col}{c.row}"].font=ARIAL
                wired+=1
    cover=wb["Cover"]
    cover["C12"]=("Capex block: five expansion lines, maintenance capex, terminal size feeding the "
                  "non-aero elasticity, depreciation from useful lives. Demo spend profile is a "
                  "working assumption.")
    cover["C12"].font=ARIAL
    wb.properties.creator=wb.properties.lastModifiedBy="Avia Solutions"
    wb.save(out)
    print(f"saved {out} | capex rows: gc={gc} maint={mt} sqm_growth={tg} dep_total={dt} | non-aero terminal rows wired: {wired}")
    return {"gc":gc,"mt":mt,"tg":tg,"dt":dt}

if __name__=="__main__":
    main(sys.argv[1],sys.argv[2],sys.argv[3])
