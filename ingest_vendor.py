"""Generator v1 increment 12: the vendor ingest front door (schema 11 v1.1:
'vendor P&L ingestion proposes the T6 mapping; the analyst confirms or edits it
before anything forecasts'). Takes whatever the vendor sent as a spreadsheet
(vendor model tab, management P&L, teaser table; xlsx or csv), and produces:
1. extracted line items with full provenance (file, sheet, row), year columns
   detected from headers, totals recognised and used for reconciliation rather
   than treated as lines;
2. t6_project_lines_proposed.tsv: the vendor chart of accounts with taxonomy
   mappings PROPOSED by an alias table and marked by confidence; unmapped lines
   carry '-' and BLOCK generation until the analyst fills them (the designed
   behaviour, already break-tested in t6_support);
3. t1_vendor_actuals.tsv: the vendor-grain actual rows, verbatim values;
4. t1_actuals_avia.tsv: Avia-grain actual rows for the actuals strip, produced
   only from high-confidence one-to-one mappings; split candidates are left to
   the analyst (flag, never fabricate);
5. ingest_report.md: what was read, what mapped at which confidence, what did
   not, and the stated-total reconciliation.
The alias table is a seed (taxonomy 03 labels plus TAS/Plovdiv line wording);
it merges with the extraction pipeline's resolve_aliases when the estates join.
Author: Avia Solutions."""
import sys, os, re, csv, datetime

ALIASES = {  # metric_code: keywords (lowercase substring match); order matters, first hit wins
 "rev_psc": ["passenger charge", "passenger service charge", "passenger fee", "psc"],
 "rev_landing": ["landing"],
 "rev_security": ["security"],
 "rev_acft_parking": ["aircraft parking", "parking fee"],
 "rev_terminal_fee": ["terminal fee", "terminal charge"],
 "rev_aero_other": ["aeronautical revenue", "aeronautical charges", "aero revenue", "traffic income"],
 "conc_dutyfree": ["duty free", "duty-free"],
 "conc_retail": ["retail"],
 "conc_fb": ["food and beverage", "food & beverage", "f&b", "catering income"],
 "rev_carpark": ["car park", "carpark", "parking income"],
 "rev_carrental": ["car rental", "car hire", "rent a car"],
 "rev_advertising": ["advertis"],
 "rev_lounge": ["lounge", "vip"],
 "rev_property": ["property", "rents", "rental income", "real estate"],
 "rev_fuel_throughput": ["fuel"],
 "rev_nonaero_other": ["other operating income", "other commercial", "other income", "sundry income"],
 "staff_costs": ["staff", "wages", "salaries", "personnel", "payroll"],
 "opex_utilities": ["utilities", "energy", "electricity"],
 "opex_rm": ["repairs", "maintenance", "r&m"],
 "opex_insurance": ["insurance"],
 "opex_rent": ["rent and rates", "rates", "lease costs"],
 "opex_marketing": ["marketing", "advertising costs", "sales & marketing"],
 "opex_cleaning": ["cleaning", "waste"],
 "opex_other": ["other operating costs", "other costs", "other expenses", "general and admin",
                "administrative"],
 "pax_total": ["passengers", "pax"],
 "atm_total": ["movements", "atms", "flights"],
}
TOTAL_WORDS = ["total", "ebitda", "gross profit", "operating profit", "sum of", "subtotal",
               "revenue total", "net result"]
YEAR_RE = re.compile(r"(?:fy)?\s*(19[9]\d|20[0-5]\d)\s*(a|f|b|e)?$", re.I)
# two-digit teaser years (FY23, '23, 23A); require an FY/apostrophe prefix or an A/F/B/E
# suffix to disambiguate from a plain number, and read as 20xx.
YEAR2_RE = re.compile(r"(?:fy\s*|')\s*([0-5]\d)\s*(?:a|f|b|e)?$|^([0-5]\d)\s*(?:a|f|b|e)$", re.I)

def year_of(v):
    if isinstance(v, (int, float)) and 1990 <= int(v) <= 2059 and int(v) == v:
        return int(v)
    if isinstance(v, str):
        s = v.strip().lower()
        m = YEAR_RE.match(s)
        if m: return int(m.group(1))
        m2 = YEAR2_RE.match(s)
        if m2: return 2000 + int(m2.group(1) or m2.group(2))
    if hasattr(v, "year"):
        return int(v.year)
    return None

def read_grid(path, sheet=None):
    """Returns list of sheets: (sheet_name, rows) where rows are lists of cell values."""
    if path.lower().endswith((".csv", ".tsv")):
        delim = "\t" if path.lower().endswith(".tsv") else ","
        with open(path, encoding="utf-8-sig") as fh:
            return [("csv", [r for r in csv.reader(fh, delimiter=delim)])]
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    names = [sheet] if sheet else wb.sheetnames
    return [(sn, [[c.value for c in row] for row in wb[sn].iter_rows()]) for sn in names]

def extract(path, sheet=None):
    """Find the year header, then every labelled row with numeric values under it."""
    lines, totals = [], []
    for sn, rows in read_grid(path, sheet):
        hdr_i, years = None, {}
        for i, row in enumerate(rows[:40]):
            ys = {j: year_of(v) for j, v in enumerate(row) if year_of(v)}
            if len(ys) >= 2:
                hdr_i, years = i, ys
                break
        if hdr_i is None:
            continue
        for i in range(hdr_i + 1, len(rows)):
            row = rows[i]
            label = next((str(v).strip() for j, v in enumerate(row)
                          if isinstance(v, str) and v.strip() and j < min(years)), None)
            if not label:
                continue
            vals = {}
            for j, y in years.items():
                v = row[j] if j < len(row) else None
                if isinstance(v, (int, float)):
                    vals[y] = float(v)
                elif isinstance(v, str) and v.strip():
                    try:
                        vals[y] = float(v.strip().replace(",", "").replace("(", "-").replace(")", ""))
                    except ValueError:
                        pass
            if not vals:
                continue
            rec = {"label": label, "values": vals, "sheet": sn, "row": i + 1}
            if any(w in label.lower() for w in TOTAL_WORDS):
                totals.append(rec)
            else:
                lines.append(rec)
    return lines, totals

def propose(label):
    """(code, confidence): high = single strong hit; low = weak or ambiguous."""
    lab = label.lower()
    hits = [(code, kw) for code, kws in ALIASES.items() for kw in kws if kw in lab]
    if not hits:
        return None, None
    codes = list(dict.fromkeys(c for c, _ in hits))
    if len(codes) == 1:
        return codes[0], "high" if len(hits[0][1]) >= 5 else "low"
    return codes[0], "low"          # ambiguous: propose the first, mark low

def main(path, unit, out_dir, sheet=None):
    os.makedirs(out_dir, exist_ok=True)
    lines, totals = extract(path, sheet)
    if not lines:
        raise SystemExit("INGEST BLOCK: no labelled year rows found; check the file or name the sheet")
    src = os.path.basename(path)
    today = datetime.date.today().strftime("%d %B %Y")
    t6, t1v, t1a, unmapped, low = [], [], [], [], []
    for k, rec in enumerate(lines, 1):
        lid = f"V{k:03d}"
        code, conf = propose(rec["label"])
        if code is None:
            mapping = "-"
            unmapped.append(rec["label"])
        else:
            basis = f"auto-proposed from alias table, {conf} confidence, CONFIRM before generation"
            mapping = f"{code}:1.0:{basis}"
            if conf == "low":
                low.append((rec["label"], code))
        t6.append(f"{lid}\t{rec['label']}\t-\t{mapping}\tthis_line\tvendor")
        # sign normalisation: vendor P&Ls carry costs negative; the model holds cost lines positive
        is_cost = bool(code) and (code.startswith("opex_") or code == "staff_costs")
        mostly_neg = sum(1 for v in rec["values"].values() if v < 0) > len(rec["values"]) / 2
        flip = is_cost and mostly_neg
        for y, v in sorted(rec["values"].items()):
            prov = f"{src}, sheet {rec['sheet']}, row {rec['row']}"
            t1v.append(f"{lid}\ttotal\t1\t{y}\t{v}\t{unit}\tactual\tlevel\t\t\t\t{prov}")
            aero_rate_input = bool(code) and code.startswith("rev_") and code not in (
                "rev_carpark","rev_carrental","rev_advertising","rev_lounge","rev_property",
                "rev_fuel_throughput","rev_nonaero_other")
            if code and conf == "high" and not aero_rate_input:
                vv = -v if flip else v
                note = "; sign normalised (vendor costs negative)" if flip else ""
                t1a.append(f"{code}\ttotal\t1\t{y}\t{vv}\t{unit}\tactual\tlevel\t\t\t\t{prov} via {lid}{note}")
    hdr6 = ("# T6 project_lines PROPOSED by vendor ingest | source: {0} | {1}\n"
            "# Unmapped lines carry '-' and BLOCK generation until the analyst maps them.\n"
            "# line_id\tlabel\tparent_line_id\ttaxonomy_codes\tforecast_grain\torigin\n").format(src, today)
    hdr1 = ("# T1 rows from vendor ingest | source: {0} | {1}\n"
            "# metric_code\tsegment\tcase_id\tyear\tvalue\tunit\ttemporality\tdriver_type"
            "\tstep_date\tstep_value\trepeat_years\tsource\n").format(src, today)
    open(os.path.join(out_dir, "t6_project_lines_proposed.tsv"), "w", encoding="utf-8").write(hdr6 + "\n".join(t6) + "\n")
    open(os.path.join(out_dir, "t1_vendor_actuals.tsv"), "w", encoding="utf-8").write(hdr1 + "\n".join(t1v) + "\n")
    open(os.path.join(out_dir, "t1_actuals_avia.tsv"), "w", encoding="utf-8").write(
        hdr1.replace("T1 rows", "T1 Avia-grain actual rows (high-confidence one-to-one maps only)") + "\n".join(t1a) + "\n")
    # reconciliation against stated totals
    rep = [f"# Vendor ingest report | {src} | {today}", "",
           f"Lines extracted: {len(lines)} across {len(set(l['sheet'] for l in lines))} sheet(s); "
           f"stated totals recognised: {len(totals)}.",
           f"Mapping: {len(lines)-len(unmapped)} proposed ({len(low)} low confidence), "
           f"{len(unmapped)} unmapped and BLOCKING: {', '.join(unmapped) if unmapped else 'none'}.", ""]
    if low:
        rep.append("Low confidence, review first: " + "; ".join(f"'{l}' to {c}" for l, c in low))
    for t in totals:
        tl = t["label"].lower()
        for y, tv in sorted(t["values"].items()):
            if "revenue" in tl or "income" in tl:
                ssum = sum(v for r in lines for yy, v in r["values"].items() if yy == y and v > 0)
                kind = "positive lines"
            elif "cost" in tl or "expense" in tl:
                ssum = sum(v for r in lines for yy, v in r["values"].items() if yy == y and v < 0)
                kind = "negative lines"
            else:
                ssum = sum(r["values"].get(y, 0) for r in lines)
                kind = "all lines net"
            rep.append(f"Stated '{t['label']}' {y}: {tv}; extracted {kind} sum {round(ssum,3)}; "
                       f"difference {round(ssum-tv,3)}.")
    rep.append("")
    rep.append("Aero revenue lines are kept at vendor grain only: the workbook's aero inputs are "
               "unit RATES, and a P&L gives revenue amounts; rate calibration (revenue over driver) "
               "is an explicit later step. Traffic totals need the dom/intl split (Sabre/OAG) before "
               "they can place.")
    rep.append("")
    rep.append("Next: review and edit t6_project_lines_proposed.tsv (the mapping screen in the Studio "
               "will replace this file edit), then t6_support.py validate blocks until every line maps. "
               "t1_actuals_avia.tsv feeds the actuals strip; split-candidate lines are left to the analyst.")
    open(os.path.join(out_dir, "ingest_report.md"), "w", encoding="utf-8").write("\n".join(rep) + "\n")
    print(f"ingested {len(lines)} lines | mapped {len(lines)-len(unmapped)} ({len(low)} low) | "
          f"unmapped {len(unmapped)} | avia-grain actual rows {len(t1a)} | report in {out_dir}")

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4] if len(sys.argv) > 4 else None)
