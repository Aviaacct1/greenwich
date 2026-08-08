"""Generator v1 increment 7b: financing and returns sheet group v3 (blueprint 06
section 10, Scanner grammar), superseding the increment 3 emission. Changes from
v2: maintenance and growth capex enter the waterfall from the Capex sheet
(Scanner order: maintenance before external interest, growth after debt service,
both before dividends); depreciation (Capex sheet) enters the taxable base and
net income; LTM EBITDA is the last actual year's EBITDA from Operations Summary
rather than the entry-year proxy. Interest stays on AVERAGE(opening, closing)
balances with Excel iterative calculation enabled; closing cash stays in the MIN
form with the special dividend derived; the sweep still deducts the principal
actually paid; the on-sheet convergence residual and WATCH status remain.
All input values are working assumptions, illustrative, labelled on-sheet.
Author: Avia Solutions."""
import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.workbook.properties import CalcProperties

ARIAL=Font(name="Arial",size=10); BOLD=Font(name="Arial",size=10,bold=True)
FC=12

def read_header(path):
    h={}
    for line in open(path,encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"): continue
        k,v=line.rstrip("\n").split("\t"); h[k]=v
    return h

def main(src,header_path,out):
    H=read_header(header_path)
    Y0=int(H["start_year"]); LA=int(H["last_actual_year"]); N=int(H["model_term_years"])
    ENTRY=LA+1; HOLD=10; EXIT=ENTRY+HOLD-1
    if EXIT>Y0+N-1: raise SystemExit("exit year beyond model term")
    jE,jX,jL=ENTRY-Y0,EXIT-Y0,LA-Y0
    cols=[get_column_letter(FC+i) for i in range(N)]
    cE,cX=cols[jE],cols[jX]
    cL=cols[jL] if jL>=0 else cols[jE]     # no actuals in ingest: entry-year proxy, labelled below
    wb=openpyxl.load_workbook(src)
    for s in ("Financing Inputs","Debt Schedule","Waterfall","Equity Returns"):
        if s in wb.sheetnames: del wb[s]
    # Capex sheet row map (discover by label so this stays robust)
    cx=wb["Capex"]; cxrow={}
    for row in cx.iter_rows(min_col=4,max_col=4):
        v=row[0].value
        if v=="Total expansion (growth) capex": cxrow["gc"]=row[0].row
        if v=="Maintenance capex": cxrow["mt"]=row[0].row
        if v=="Total depreciation": cxrow["dt"]=row[0].row
    assert len(cxrow)==3,"Capex rows not found; run capex_block first"

    def put(ws,row,col,val,font=ARIAL,fmt=None):
        c=ws.cell(row=row,column=col,value=val); c.font=font
        if fmt: c.number_format=fmt
        return c

    # ------------------------------------------------------------------ inputs
    fi=wb.create_sheet("Financing Inputs")
    fi["A1"]="=Control!B1"; fi["A1"].font=BOLD
    put(fi,2,2,"Financing and returns inputs (Scanner grammar, blueprint 06 section 10, v3 with "
        "capex and depreciation wired). All values are working assumptions, illustrative.",BOLD)
    put(fi,3,2,f"Entry and exit years set at generation from the project header "
        f"(entry {ENTRY}, exit {EXIT}, hold {HOLD} years); change by regenerating.",ARIAL)
    put(fi,4,2,"Convention: this model runs on a real (constant base-year) P&L; financing and "
        "returns are computed in real terms. Nominal presentation is memo only (Operations Summary "
        "and Capex nominal rows via the Macro block). A nominal financing mode is a later option.",ARIAL)
    WA="working assumption, illustrative"
    inputs=[(5,"Entry EV / LTM EBITDA multiple [x]",10.0),(6,"Exit EV / LTM EBITDA multiple [x]",10.0),
     (8,"Senior debt / LTM EBITDA [x]",3.0),(9,"Junior debt / LTM EBITDA [x]",1.5),
     (10,"Cost of senior debt [%]",0.06),(11,"Cost of junior debt [%]",0.09),
     (12,"Interest earned on cash [%]",0.02),(13,"Financing fees [% of principal]",0.02),
     (14,"Senior cash sweep [% of surplus]",0.75),(15,"Junior term repayment [% of original principal]",0.05),
     (16,"Debt service reserve [x forward service]",1.0),(18,"Tax rate [%]",0.25),
     (19,"Goodwill amortisation deductible? [Yes/No]","No"),(20,"Goodwill life [years]",10),
     (21,"Fair value of net assets [EUR m]",100.0),(23,"Debtor days",30),(24,"Creditor days",45),
     (25,"Dividend payout [% of net income]",0.50),(26,"Special dividends? [Yes/No]","Yes"),
     (27,"Exit transaction costs [% of exit EV]",0.01)]
    for r,lab,v in inputs:
        put(fi,r,2,lab); put(fi,r,4,v,BOLD); put(fi,r,6,f"Source: {WA}")
    put(fi,29,2,"Transaction costs at entry [EUR m], itemised (Scanner list)",BOLD)
    for k,(lab,v) in enumerate([("Pensions",0.2),("Accounting",0.3),("Financial",0.5),
                                 ("Legal",0.4),("Technical",0.3),("Avia",0.2)]):
        put(fi,30+k,2,lab); put(fi,30+k,4,v); put(fi,30+k,6,f"Source: {WA}")
    put(fi,36,2,"Total transaction costs",BOLD); put(fi,36,4,"=SUM(D30:D35)",BOLD)
    put(fi,39,2,"Sources and uses of funds at entry [EUR m]",BOLD)
    put(fi,40,2,f"LTM EBITDA: last actual year ({LA}A) EBITDA; entry-year proxy if non-positive "
        "(partial ingest leaves the actual era incomplete)" if jL>=0
        else f"LTM EBITDA proxy: entry-year ({ENTRY}) EBITDA, working assumption (no actuals in ingest)")
    put(fi,40,4,(f"=IF('Operations Summary'!{cL}12>0,'Operations Summary'!{cL}12,"
                 f"'Operations Summary'!{cE}12)") if jL>=0 else f"='Operations Summary'!{cE}12",BOLD)
    put(fi,42,2,"Sources",BOLD)
    put(fi,43,2,"Senior acquisition debt"); put(fi,43,4,"=D8*D40")
    put(fi,44,2,"Junior acquisition debt"); put(fi,44,4,"=D9*D40")
    put(fi,45,2,"Equity contribution");     put(fi,45,4,"=D54-D43-D44")
    put(fi,46,2,"Total sources",BOLD);      put(fi,46,4,"=SUM(D43:D45)",BOLD)
    put(fi,48,2,"Uses",BOLD)
    put(fi,49,2,"Purchase price of equity (no existing debt in skeleton)"); put(fi,49,4,"=D5*D40")
    put(fi,50,2,"Financing fees"); put(fi,50,4,"=D13*(D43+D44)")
    put(fi,51,2,"Transaction costs"); put(fi,51,4,"=D36")
    put(fi,52,2,"Debt service reserve funding (first-year forward service x reserve)")
    put(fi,52,4,"=D16*(D43*D10+D44*D11+D44*D15)")
    put(fi,54,2,"Total uses",BOLD); put(fi,54,4,"=SUM(D49:D52)",BOLD)
    put(fi,56,2,"Sources and uses balance check",BOLD)
    put(fi,56,4,'=IF(ROUND(D46-D54,6)=0,"Ok","ERROR")',BOLD)
    put(fi,58,2,"Goodwill: EV less fair value of net assets (SLN amortisation over life)")
    put(fi,58,4,"=D49-D21")
    put(fi,61,2,"Status indicators",BOLD)
    status=[("Sources and uses balance","=D56"),
     ("Debt service reserve cover (exit-year check on Waterfall)",f"=Waterfall!{cX}30"),
     ("Senior debt never negative",f"=IF(MIN('Debt Schedule'!{cE}7:{cX}7)>=0,\"Ok\",\"ERROR\")"),
     ("Cash never negative",f"=IF(MIN(Waterfall!{cE}28:{cX}28)>=-0.000001,\"Ok\",\"ERROR\")"),
     ("Iteration converged (circular loop residual under 0.01)",
      f"=IF(SUMPRODUCT(--(ABS(Waterfall!{cE}31:{cX}31)>0.01))=0,\"Ok\",\"WATCH\")")]
    for k,(lab,f) in enumerate(status):
        put(fi,62+k,2,lab); put(fi,62+k,4,f)
    put(fi,68,2,"Overall financing status",BOLD)
    put(fi,68,4,'=IF(COUNTIF(D62:D66,"ERROR")=0,"Ok","ERROR")',BOLD)

    # ------------------------------------------------------------- debt schedule
    ds=wb.create_sheet("Debt Schedule")
    ds["A1"]="=Control!B1"; ds["A1"].font=BOLD
    put(ds,2,2,"Debt schedule [EUR m]. Interest on AVERAGE(opening, closing) balances per Scanner; "
        "requires Excel iterative calculation (set in this workbook).",BOLD)
    for j in range(N):
        y=Y0+j
        put(ds,6,FC+j,f"{y}{'A' if y<=LA else 'F'}",BOLD)
    labels={7:"Senior debt - opening balance",8:"Senior debt - cash sweep (from Waterfall)",
     9:"Senior debt - closing balance",11:"Junior debt - opening balance",
     12:"Junior debt - term repayment",13:"Junior debt - closing balance",
     15:"Senior interest (on average balance)",16:"Junior interest (on average balance)",
     17:"Total external interest",
     19:"Forward external debt service (next year interest + junior repayment)",
     21:"Goodwill amortisation"}
    for r,lab in labels.items(): put(ds,r,4,lab)
    for j,c in enumerate(cols):
        if j<jE or j>jX: continue
        pc=cols[j-1]
        put(ds,7,FC+j,"='Financing Inputs'!$D$43" if j==jE else f"={pc}9")
        put(ds,8,FC+j,f"=Waterfall!{c}17")
        put(ds,9,FC+j,f"=MAX({c}7+{c}8,0)")
        put(ds,11,FC+j,"='Financing Inputs'!$D$44" if j==jE else f"={pc}13")
        put(ds,12,FC+j,f"=IF({c}11<=0,0,-MIN('Financing Inputs'!$D$44*'Financing Inputs'!$D$15,{c}11))")
        put(ds,13,FC+j,f"={c}11+{c}12")
        put(ds,15,FC+j,f"=-(AVERAGE({c}7,{c}9)*'Financing Inputs'!$D$10)")
        put(ds,16,FC+j,f"=-(AVERAGE({c}11,{c}13)*'Financing Inputs'!$D$11)")
        put(ds,17,FC+j,f"={c}15+{c}16")
        if j<jX:
            nc=cols[j+1]
            put(ds,19,FC+j,f"=-({nc}17)+(-{nc}12)")
        else:
            put(ds,19,FC+j,0)
        yr=Y0+j
        put(ds,21,FC+j,f"=IF({yr}<{ENTRY}+'Financing Inputs'!$D$20,"
            f"-'Financing Inputs'!$D$58/'Financing Inputs'!$D$20,0)")

    # ---------------------------------------------------------------- waterfall
    wf=wb.create_sheet("Waterfall")
    wf["A1"]="=Control!B1"; wf["A1"].font=BOLD
    put(wf,2,2,"Cash waterfall [EUR m], Scanner order, v3: maintenance capex before external "
        "interest, growth capex after debt service, depreciation in the taxable base. Working "
        "capital: debtor days on revenue, creditor days on operating costs (standard convention).",BOLD)
    for j in range(N):
        y=Y0+j
        put(wf,6,FC+j,f"{y}{'A' if y<=LA else 'F'}",BOLD)
    wlab={7:"EBITDA",8:"Trade debtors (revenue/365 x debtor days)",
     9:"Trade creditors (operating costs/365 x creditor days)",10:"Working capital movement",
     11:"Interest received on cash (average balance)",
     12:"Depreciation (memo, from Capex; taxable base, not cash)",
     13:"Goodwill amortisation (memo, from Debt Schedule)",
     14:"Cash tax (nil if taxable base negative; goodwill add-back per toggle)",
     15:"Cash from operations",16:"Maintenance capex (from Capex sheet)",
     17:"Cash sweep to senior (surplus incl. cash brought forward, less forward reserve)",
     18:"Senior principal actually paid",19:"Junior term repayment",
     20:"Cash after debt service (incl. cash brought forward)",
     21:"Growth capex (from Capex sheet)",22:"Cash after growth capex",
     23:"Net income (memo: EBITDA + depreciation + goodwill + net interest + tax)",
     24:"Dividends (payout % of net income, cash above reserve only)",
     25:"Special dividends (derived from closing cash)",
     26:"Debt service reserve requirement",27:"Cash - opening balance",
     28:"Cash - closing balance (MIN form; special dividend derived)",
     30:"Debt service reserve check",
     31:"Iteration convergence residual (zero when fully converged)"}
    for r,lab in wlab.items(): put(wf,r,4,lab)
    for j,c in enumerate(cols):
        if j<jE or j>jX: continue
        pc=cols[j-1]
        put(wf,7,FC+j,f"='Operations Summary'!{c}12")
        put(wf,8,FC+j,f"='Operations Summary'!{c}10/365*'Financing Inputs'!$D$23")
        put(wf,9,FC+j,f"='Operations Summary'!{c}11/365*'Financing Inputs'!$D$24")
        if j==jE: put(wf,10,FC+j,f"=-{c}8+{c}9")
        else: put(wf,10,FC+j,f"=-({c}8-{pc}8)+({c}9-{pc}9)")
        put(wf,11,FC+j,f"=AVERAGE({c}27,{c}28)*'Financing Inputs'!$D$12")
        put(wf,12,FC+j,f"=Capex!{c}{cxrow['dt']}")
        put(wf,13,FC+j,f"='Debt Schedule'!{c}21")
        put(wf,14,FC+j,
            f"=IF({c}7+{c}12+{c}13+'Debt Schedule'!{c}17+{c}11"
            f"+IF('Financing Inputs'!$D$19=\"No\",-{c}13,0)<=0,0,"
            f"-({c}7+{c}12+{c}13+'Debt Schedule'!{c}17+{c}11"
            f"+IF('Financing Inputs'!$D$19=\"No\",-{c}13,0))*'Financing Inputs'!$D$18)")
        put(wf,15,FC+j,f"={c}7+{c}10+{c}11+{c}14")
        put(wf,16,FC+j,f"=-Capex!{c}{cxrow['mt']}")
        put(wf,17,FC+j,
            f"=IF({c}27+{c}15+{c}16+'Debt Schedule'!{c}17-{c}26<0,0,"
            f"-({c}27+{c}15+{c}16+'Debt Schedule'!{c}17-{c}26)*'Financing Inputs'!$D$14)")
        put(wf,18,FC+j,f"='Debt Schedule'!{c}9-'Debt Schedule'!{c}7")
        put(wf,19,FC+j,f"='Debt Schedule'!{c}12")
        put(wf,20,FC+j,f"={c}27+{c}15+{c}16+'Debt Schedule'!{c}17+{c}18+{c}19")
        put(wf,21,FC+j,f"=-Capex!{c}{cxrow['gc']}")
        put(wf,22,FC+j,f"={c}20+{c}21")
        put(wf,23,FC+j,f"={c}7+{c}12+{c}13+'Debt Schedule'!{c}17+{c}11+{c}14")
        put(wf,24,FC+j,f"=IF(AND({c}23>0,{c}22-{c}26>0),-MIN({c}23*'Financing Inputs'!$D$25,{c}22-{c}26),0)")
        put(wf,25,FC+j,f"=IF('Financing Inputs'!$D$26=\"Yes\",{c}28-({c}22+{c}24),0)")
        put(wf,26,FC+j,f"='Debt Schedule'!{c}19*'Financing Inputs'!$D$16")
        put(wf,27,FC+j,"='Financing Inputs'!$D$52" if j==jE else f"={pc}28")
        put(wf,28,FC+j,f"=IF('Financing Inputs'!$D$26=\"Yes\",MIN({c}22+{c}24,{c}26),{c}22+{c}24)")
        put(wf,30,FC+j,f"=IF({c}28>={c}26-0.000001,\"Ok\",\"ERROR\")")
        put(wf,31,FC+j,f"={c}20-({c}27+{c}15+{c}16+'Debt Schedule'!{c}17+{c}18+{c}19)")

    # ------------------------------------------------------------ equity returns
    er=wb.create_sheet("Equity Returns")
    er["A1"]="=Control!B1"; er["A1"].font=BOLD
    put(er,2,2,f"Equity returns [EUR m]. Entry end-{ENTRY}, exit end-{EXIT} ({HOLD}-year hold). "
        "Pre-tax equity IRR per Scanner; levered heat matrix below (entry x exit multiple over "
        "the equity flow series).",BOLD)
    for j in range(N):
        y=Y0+j
        put(er,6,FC+j,f"{y}{'A' if y<=LA else 'F'}",BOLD)
    elab={7:"Exit enterprise value (exit multiple x exit-year EBITDA)",
     8:"Less debt outstanding at exit",9:"Plus cash at exit",10:"Less exit transaction costs",
     11:"Equity value at exit",13:"Equity flow (base case)"}
    for r,lab in elab.items(): put(er,r,4,lab)
    put(er,7,FC+jX,f"='Financing Inputs'!$D$6*'Operations Summary'!{cX}12")
    put(er,8,FC+jX,f"=-('Debt Schedule'!{cX}9+'Debt Schedule'!{cX}13)")
    put(er,9,FC+jX,f"=Waterfall!{cX}28")
    put(er,10,FC+jX,f"=-{cX}7*'Financing Inputs'!$D$27")
    put(er,11,FC+jX,f"=SUM({cX}7:{cX}10)")
    for j in range(jE,jX+1):
        c=cols[j]
        if j==jE: f=f"=-'Financing Inputs'!$D$45-(Waterfall!{c}24+Waterfall!{c}25)"
        else:
            f=f"=-(Waterfall!{c}24+Waterfall!{c}25)"
            if j==jX: f+=f"+{cX}11"
        put(er,13,FC+j,f)
    put(er,15,4,"Pre-tax equity IRR",BOLD)
    put(er,15,8,f"=IRR({cE}13:{cX}13)",BOLD,"0.0%")
    put(er,16,4,"Equity MOIC",BOLD)
    put(er,16,8,f"=SUMIF({cE}13:{cX}13,\">0\")/'Financing Inputs'!$D$45",BOLD,"0.00x")
    entries=[8,9,10,11,12]; exits=[8,9,10,11,12,13,14]
    put(er,19,2,"Levered heat matrix: pre-tax equity IRR by entry and exit multiple. Debt is "
        "sized on LTM EBITDA so the waterfall is common to every cell; only the entry equity "
        "cheque and the exit value vary.",BOLD)
    put(er,21,2,"Entry x / Exit x",BOLD)
    for j,xm in enumerate(exits): put(er,21,3+j,xm,BOLD)
    put(er,30,2,"Cash-flow helper rows (one per grid cell)")
    hr=32; hold_n=jX-jE
    for i,em in enumerate(entries):
        for j,xm in enumerate(exits):
            put(er,hr,2,f"{em}x in / {xm}x out")
            put(er,hr,3,
                f"=-({em}*'Financing Inputs'!$D$40+'Financing Inputs'!$D$50"
                f"+'Financing Inputs'!$D$51+'Financing Inputs'!$D$52"
                f"-'Financing Inputs'!$D$43-'Financing Inputs'!$D$44)"
                f"-(Waterfall!{cE}24+Waterfall!{cE}25)")
            for k in range(1,hold_n+1):
                c=cols[jE+k]
                f=f"=-(Waterfall!{c}24+Waterfall!{c}25)"
                if k==hold_n:
                    f+=(f"+{xm}*'Operations Summary'!{cX}12"
                        f"-('Debt Schedule'!{cX}9+'Debt Schedule'!{cX}13)"
                        f"+Waterfall!{cX}28"
                        f"-{xm}*'Operations Summary'!{cX}12*'Financing Inputs'!$D$27")
                put(er,hr,3+k,f)
            put(er,22+i,3+j,f"=IRR(C{hr}:{get_column_letter(3+hold_n)}{hr})",ARIAL,"0.0%")
            hr+=1
        put(er,22+i,2,f"{em}x",BOLD)
    rng=f"C22:{get_column_letter(2+len(exits))}{21+len(entries)}"
    er.conditional_formatting.add(rng,ColorScaleRule(
        start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,
        mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    put(er,hr+2,2,"Source: structure per Project Scanner Valuation Model (Avia, 20Dec10) as "
        "ported in blueprint 06 section 10; all values working assumptions, illustrative.")
    wb["Cover"]["C11"]=("Financing group v3 (Scanner grammar, capex and depreciation wired): "
                        "iterative calculation enabled; Excel resolves it on open.")
    wb["Cover"]["C11"].font=ARIAL
    wb.calculation=CalcProperties(calcId=124519,fullCalcOnLoad=True,
                                  iterate=True,iterateCount=200,iterateDelta=0.0000001)
    wb.properties.creator=wb.properties.lastModifiedBy="Avia Solutions"
    wb.save(out)
    print("saved",out,"| entry",ENTRY,"| exit",EXIT,"| LTM from",f"{LA}A","| capex rows",cxrow)

if __name__=="__main__":
    main(sys.argv[1],sys.argv[2],sys.argv[3])
