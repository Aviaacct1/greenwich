"""Generator v1 increment 9: output suite reinstated on the current chain
(v0.2's Per Pax sheet, native chart set and valuation heat matrix, lost when
the incremental chain rebuilt from v0). Changes from v0.2: year spine and A/F
labels from the project header; charts carry unit and period in the title and
a source line on the axis title (house chart rules); the unlevered heat matrix
now runs on EBITDA LESS maintenance and growth capex (the Capex sheet exists),
with the same entry/exit years as the financing group; the old checks block is
NOT re-emitted (the Checks sheet owns that). Author: Avia Solutions."""
import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

ARIAL=Font(name="Arial",size=10); BOLD=Font(name="Arial",size=10,bold=True)
FC=12

def read_header(p):
    h={}
    for line in open(p,encoding="utf-8"):
        if line.startswith("#") or not line.strip() or line.startswith("key"): continue
        k,v=line.rstrip("\n").split("\t"); h[k]=v
    return h

def main(src,hdr,out):
    H=read_header(hdr); Y0=int(H["start_year"]); LA=int(H["last_actual_year"]); N=int(H["model_term_years"])
    ENTRY=LA+1; HOLD=10; EXIT=ENTRY+HOLD-1
    jE,jX=ENTRY-Y0,EXIT-Y0
    cols=[get_column_letter(FC+i) for i in range(N)]
    span=f"{Y0}-{Y0+N-1}, actuals to {LA} then forecast" if LA>=Y0 else f"{Y0}-{Y0+N-1}, forecast"
    wb=openpyxl.load_workbook(src)
    for s in ("Per Pax","Charts","Valuation"):
        if s in wb.sheetnames: del wb[s]
    osum=wb["Operations Summary"]; cx=wb["Capex"]
    cxrow={}
    for row in cx.iter_rows(min_col=4,max_col=4):
        v=row[0].value
        if v=="Total expansion (growth) capex": cxrow["gc"]=row[0].row
        if v=="Maintenance capex": cxrow["mt"]=row[0].row
    OS_LINES={8:"Aeronautical revenue",9:"Non-aeronautical revenue",10:"Total revenue",
              11:"Operating costs",12:"EBITDA"}
    # Per Pax
    pp=wb.create_sheet("Per Pax")
    pp["A1"]="=Control!B1"; pp["A1"].font=BOLD
    pp["B2"]="Per-passenger metrics [EUR per pax] (line EUR m / total pax m). Source: model."
    pp["B2"].font=BOLD
    for j in range(N):
        y=Y0+j
        pp.cell(row=6,column=FC+j,value=f"{y}{'A' if y<=LA else 'F'}").font=BOLD
    for k,(r,lbl) in enumerate(OS_LINES.items()):
        pr=8+k
        pp.cell(row=pr,column=4,value=f"{lbl} per pax").font=ARIAL
        for col in cols:
            pp[f"{col}{pr}"]=f"=IFERROR('Operations Summary'!{col}{r}/Aero!{col}11,0)"
            pp[f"{col}{pr}"].font=ARIAL
    # Charts
    ch_ws=wb.create_sheet("Charts")
    ch_ws["A1"]="Avia curve suite: every Operations Summary line, total and per pax. House format; native Excel charts are the review layer, the matplotlib factory (WS11) the report layer."
    ch_ws["A1"].font=BOLD
    def add_chart(sheetname,datarow,title,unit,anchor):
        ch=LineChart(); ch.title=f"{title} [{unit}], {span}"
        ch.style=2; ch.height=6.5; ch.width=13
        ch.y_axis.title=unit
        ch.x_axis.title="Source: AviaSolutions analysis (illustrative demo inputs)"
        ws=wb[sheetname]
        data=Reference(ws,min_col=FC,max_col=FC+N-1,min_row=datarow,max_row=datarow)
        cats=Reference(osum,min_col=FC,max_col=FC+N-1,min_row=6,max_row=6)
        ch.add_data(data,titles_from_data=False); ch.set_categories(cats)
        ch.series[0].smooth=False
        ch_ws.add_chart(ch,anchor)
    ar=3
    for k,(r,lbl) in enumerate(OS_LINES.items()):
        add_chart("Operations Summary",r,lbl,"EUR m",f"B{ar}")
        add_chart("Per Pax",8+k,f"{lbl} per pax","EUR per pax",f"L{ar}")
        ar+=14
    # Valuation heat matrix (unlevered, EBITDA less capex)
    val=wb.create_sheet("Valuation")
    val["A1"]="=Control!B1"; val["A1"].font=BOLD
    val["B2"]="Valuation heat matrix: pre-tax UNLEVERED IRR by entry and exit multiple."
    val["B2"].font=BOLD
    val["B3"]=(f"Entry end-{ENTRY}, exit end-{EXIT}; cash flows = EBITDA less maintenance and "
               "growth capex (Operations Summary and Capex sheets); the levered variant is on "
               "Equity Returns. Source: model; multiples working assumptions.")
    val["B3"].font=ARIAL
    entries=[8,9,10,11,12]; exits=[8,9,10,11,12,13,14]
    val["B5"]="Entry x / Exit x"; val["B5"].font=BOLD
    for j,xm in enumerate(exits): val.cell(row=5,column=3+j,value=xm).font=BOLD
    hold_cols=cols[jE+1:jX+1]
    hr=30
    val["B28"]="Cash-flow helper rows (one per grid cell)"; val["B28"].font=ARIAL
    cE=cols[jE]; cXl=cols[jX]
    ucf=lambda c: f"('Operations Summary'!{c}12-Capex!{c}{cxrow['mt']}-Capex!{c}{cxrow['gc']})"
    for i,em in enumerate(entries):
        for j,xm in enumerate(exits):
            val.cell(row=hr,column=2,value=f"{em}x in / {xm}x out").font=ARIAL
            val.cell(row=hr,column=3,value=f"=-{em}*'Operations Summary'!{cE}12").font=ARIAL
            for k,col in enumerate(hold_cols):
                f=f"={ucf(col)}" if k<len(hold_cols)-1 else f"={ucf(col)}+{xm}*'Operations Summary'!{cXl}12"
                val.cell(row=hr,column=4+k,value=f).font=ARIAL
            c=val.cell(row=7+i,column=3+j,value=f"=IRR(C{hr}:{get_column_letter(3+len(hold_cols))}{hr})")
            c.number_format="0.0%"; c.font=ARIAL
            hr+=1
        val.cell(row=7+i,column=2,value=f"{em}x").font=BOLD
    rng=f"C7:{get_column_letter(2+len(exits))}{6+len(entries)}"
    val.conditional_formatting.add(rng,ColorScaleRule(
        start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,
        mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    wb["Cover"]["C13"]="Output suite: Per Pax, Charts (house format), unlevered heat matrix on Valuation; levered on Equity Returns; rules on Checks."
    wb["Cover"]["C13"].font=ARIAL
    wb.properties.creator=wb.properties.lastModifiedBy="Avia Solutions"
    wb.save(out)
    print("saved",out,"| charts:",len(ch_ws._charts),"| unlevered matrix entry",ENTRY,"exit",EXIT)

if __name__=="__main__":
    main(sys.argv[1],sys.argv[2],sys.argv[3])
