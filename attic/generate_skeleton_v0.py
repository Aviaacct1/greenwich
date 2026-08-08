"""
Avia standard airport business-plan model: skeleton generator v0.
Implements the grammars in '06 WS1 Blueprint - Standard Model Skeleton - 15 July 2026'.
v0 scope: Cover, Control, Inputs (case blocks + INDEX selectors), Aero (worked block),
Non-aero (elasticity block per category), Opex (staff + categories), Returns (stub).
Author: Avia Solutions.
"""
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

START_YEAR, END_YEAR = 2024, 2055
FIRST_YCOL = 12  # column L
NYEARS = END_YEAR - START_YEAR + 1
CASES = ["Base", "Spare 1", "Spare 2", "Spare 3", "Spare 4"]
ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)

AERO_LINES = [  # (label, metric_code, driver)
    ("Passenger service charge", "rev_psc", "pax_total"),
    ("Landing fee - domestic", "rev_landing", "atm_dom"),
    ("Landing fee - international", "rev_landing", "atm_intl"),
    ("Landing fee - cargo", "rev_landing", "atm_cargo"),
    ("Security charge", "rev_security", "pax_total"),
    ("Aircraft parking charge", "rev_acft_parking", "atm_total"),
    ("Night-time surcharge", "rev_night_surcharge", "atm_total"),
    ("Terminal fee", "rev_terminal_fee", "pax_total"),
    ("CUTE / IT charge", "rev_cute", "pax_total"),
    ("PBB / airbridge charge", "rev_pbb", "atm_total"),
    ("Other aeronautical", "rev_aero_other", "pax_total"),
]
NONAERO_CATS = [
    ("Duty free", "conc_dutyfree"), ("Specialty retail", "conc_retail"),
    ("Food & beverage", "conc_fb"), ("Advertising", "rev_advertising"),
    ("Car parking", "rev_carpark"), ("Car rental", "rev_carrental"),
    ("Lounge", "rev_lounge"), ("Property rental", "rev_property"),
    ("Fuel throughput", "rev_fuel_throughput"), ("Other non-aero", "rev_nonaero_other"),
]
OPEX_CATS = [
    ("Staff costs", "staff_costs"), ("Utilities", "opex_utilities"),
    ("Repairs and maintenance", "opex_rm"), ("Insurance", "opex_insurance"),
    ("Rent and rates", "opex_rent"), ("Marketing", "opex_marketing"),
    ("Cleaning", "opex_cleaning"), ("Other opex", "opex_other"),
]
TRAFFIC = [("Passengers - domestic", "pax_dom", "[m pax]"),
           ("Passengers - international", "pax_intl", "[m pax]")]
ATMS = [("ATMs - domestic", "atm_dom", "[k ATMs]"), ("ATMs - international", "atm_intl", "[k ATMs]"),
        ("ATMs - cargo", "atm_cargo", "[k ATMs]"), ("ATMs - other", "atm_other", "[k ATMs]")]

def style_row(ws, row, upto=60):
    for c in range(1, upto):
        ws.cell(row=row, column=c).font = ARIAL

class Registry:  # code -> Inputs selector row
    def __init__(self): self.sel = {}

def year_cols():
    return [get_column_letter(FIRST_YCOL + i) for i in range(NYEARS)]

def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws["C6"] = "Project [Codename]"; ws["C6"].font = Font(name="Arial", size=20, bold=True)
    ws["C8"] = "Financial Model - skeleton v0 (generated)"; ws["C8"].font = BOLD
    ws["C10"] = "Strictly Private and Confidential"; ws["C10"].font = ARIAL
    ws["C12"] = ("This publication provides general information and should not be relied upon in "
                 "substitution for the exercise of independent judgment. Avia accepts no liability "
                 "of any kind for loss arising from the use of the material presented.")
    ws["C12"].font = ARIAL; ws["C12"].alignment = Alignment(wrap_text=True)
    ws["C14"] = "Copyright Avia Solutions Limited. All rights reserved."; ws["C14"].font = ARIAL
    return ws

def build_control(wb):
    ws = wb.create_sheet("Control")
    ws["B1"] = "[Project]"; ws["B1"].font = BOLD
    for i, cl in enumerate(CASES):
        ws.cell(row=6, column=12 + i, value=cl).font = BOLD  # L6:P6
    ws["J6"] = 1  # running case index
    ws["B2"] = '="Running Case: "&INDEX($L$6:$P$6,$J$6)'
    ws["D4"] = "Case Choice (1-5) is set per line in column K below"; ws["D4"].font = ARIAL
    return ws

def input_block(inp, ctrl, reg, label, code, unit, ctrl_row, base_value=0):
    """Emit five case rows + INDEX selector on Inputs; case-choice cell on Control col K."""
    r0 = inp.max_row + 2
    ctrl.cell(row=ctrl_row, column=7, value=label).font = ARIAL         # Control!G label
    ctrl.cell(row=ctrl_row, column=11, value=1).font = ARIAL            # Control!K case choice
    inp.cell(row=r0, column=4, value=f"=Control!$G${ctrl_row}").font = BOLD
    for i in range(5):
        r = r0 + 1 + i
        lab = "Base case" if i == 0 else f"=Control!${get_column_letter(13+i-1)}$6"
        inp.cell(row=r, column=6, value=lab).font = ARIAL
        inp.cell(row=r, column=7, value=unit).font = ARIAL
        for j, col in enumerate(year_cols()):
            cell = inp[f"{col}{r}"]
            if i == 0:
                cell.value = base_value if j == 0 else f"={col}{r}"  # placeholder flat line
                if j > 0: cell.value = f"={get_column_letter(FIRST_YCOL+j-1)}{r}"
            else:
                cell.value = f"={col}{r-1}"   # spare defaults to row above (Plovdiv convention)
            cell.font = ARIAL
    rs = r0 + 6
    inp.cell(row=rs, column=6, value=f"=F{r0+1}").font = BOLD
    inp.cell(row=rs, column=7, value=unit).font = ARIAL
    inp.cell(row=rs, column=9, value=f"=Control!$K${ctrl_row}").font = ARIAL  # I col: case no.
    for col in year_cols():
        inp[f"{col}{rs}"] = f"=INDEX({col}{r0+1}:{col}{r0+5},$I{rs})"
        inp[f"{col}{rs}"].font = ARIAL
    reg.sel[code] = rs
    return rs

def header_block(ws, reg):
    ws["A1"] = "=Control!B1"; ws["A1"].font = BOLD
    ws["A2"] = "=Control!B2"; ws["A2"].font = ARIAL
    for j in range(NYEARS):
        ws.cell(row=6, column=FIRST_YCOL + j, value=START_YEAR + j).font = BOLD
    ws["D8"] = "Traffic (m pax, two-way)"; ws["D8"].font = BOLD
    for k, (lbl, code, _) in enumerate(TRAFFIC):
        r = 9 + k
        ws.cell(row=r, column=4, value=lbl).font = ARIAL
        for col in year_cols():
            ws[f"{col}{r}"] = f"=Inputs!{col}{reg.sel[code]}"; ws[f"{col}{r}"].font = ARIAL
    ws["D11"] = "Total"; ws["D11"].font = BOLD
    for col in year_cols(): ws[f"{col}11"] = f"=SUM({col}9:{col}10)"; ws[f"{col}11"].font = ARIAL
    ws["D13"] = "ATMs (k, two-way)"; ws["D13"].font = BOLD
    for k, (lbl, code, _) in enumerate(ATMS):
        r = 14 + k
        ws.cell(row=r, column=4, value=lbl).font = ARIAL
        for col in year_cols():
            ws[f"{col}{r}"] = f"=Inputs!{col}{reg.sel[code]}"; ws[f"{col}{r}"].font = ARIAL
    ws["D18"] = "Total ATMs"; ws["D18"].font = BOLD
    for col in year_cols(): ws[f"{col}18"] = f"=SUM({col}14:{col}17)"; ws[f"{col}18"].font = ARIAL
    return 20  # next free row

DRIVER_ROW = {"pax_total": 11, "atm_dom": 14, "atm_intl": 15, "atm_cargo": 16, "atm_total": 18}

def build_aero(wb, reg):
    ws = wb.create_sheet("Aero"); r = header_block(ws, reg)
    ws.cell(row=r, column=3, value="Aeronautical revenue").font = BOLD; r += 2
    first, last = None, None
    for lbl, code, drv in AERO_LINES:
        ws.cell(row=r, column=4, value=lbl).font = ARIAL
        ws.cell(row=r, column=7, value="[EUR m]").font = ARIAL
        sel = reg.sel[f"charge:{code}:{lbl}"]
        for col in year_cols():
            ws[f"{col}{r}"] = f"={col}{DRIVER_ROW[drv]}*Inputs!{col}{sel}"
            ws[f"{col}{r}"].font = ARIAL
        first = first or r; last = r; r += 1
    ws.cell(row=r + 1, column=4, value="Total aeronautical revenue").font = BOLD
    for col in year_cols():
        ws[f"{col}{r+1}"] = f"=SUM({col}{first}:{col}{last})"; ws[f"{col}{r+1}"].font = ARIAL
    return ("Aero", r + 1)

def build_nonaero(wb, reg):
    ws = wb.create_sheet("Non-aero"); r = header_block(ws, reg)
    ws.cell(row=r, column=3, value="Non-aeronautical revenue").font = BOLD; r += 2
    totals = []
    for lbl, code in NONAERO_CATS:
        ws.cell(row=r, column=3, value=lbl).font = BOLD
        rows = {"paxg": r+1, "epax": r+2, "gpax": r+3, "term": r+4, "eterm": r+5,
                "gterm": r+6, "uplift": r+7, "tot": r+8, "rev": r+9}
        labels = ["Passenger growth", "Elasticity to passenger growth", "Growth from elasticity to traffic",
                  "Terminal size growth", "Elasticity to terminal size growth", "Growth from elasticity terminal size",
                  "Uplift", "Total uplift", "Revenue"]
        for i, l in enumerate(labels):
            ws.cell(row=r+1+i, column=4, value=l).font = ARIAL
        cols = year_cols()
        for j, col in enumerate(cols):
            pc = cols[j-1] if j else None
            ws[f"{col}{rows['paxg']}"] = 0 if j == 0 else f"=IFERROR({col}11/{pc}11-1,0)"
            ws[f"{col}{rows['epax']}"] = f"=Inputs!{col}{reg.sel[f'el_pax:{code}']}"
            ws[f"{col}{rows['gpax']}"] = f"={col}{rows['paxg']}*{col}{rows['epax']}"
            ws[f"{col}{rows['term']}"] = 0   # placeholder: wire to Capex terminal line in v1
            ws[f"{col}{rows['eterm']}"] = f"=Inputs!{col}{reg.sel[f'el_term:{code}']}"
            ws[f"{col}{rows['gterm']}"] = f"={col}{rows['term']}*{col}{rows['eterm']}"
            ws[f"{col}{rows['uplift']}"] = f"=Inputs!{col}{reg.sel[f'uplift:{code}']}"
            ws[f"{col}{rows['tot']}"] = (f"=(1+{col}{rows['gpax']})*(1+{col}{rows['gterm']})"
                                          f"*(1+{col}{rows['uplift']})-1")
            if j == 0:
                ws[f"{col}{rows['rev']}"] = f"=Inputs!{col}{reg.sel[f'baserev:{code}']}"
            else:
                ws[f"{col}{rows['rev']}"] = f"={pc}{rows['rev']}*(1+{col}{rows['tot']})"
            for k in rows.values(): ws[f"{col}{k}"].font = ARIAL
        totals.append(rows["rev"]); r = rows["rev"] + 2
    ws.cell(row=r, column=4, value="Total non-aeronautical revenue").font = BOLD
    for col in year_cols():
        ws[f"{col}{r}"] = "=" + "+".join(f"{col}{t}" for t in totals); ws[f"{col}{r}"].font = ARIAL
    return ("Non-aero", r)

def build_opex(wb, reg):
    ws = wb.create_sheet("Opex"); r = header_block(ws, reg)
    ws.cell(row=r, column=3, value="Operating costs").font = BOLD; r += 2
    first, last = None, None
    for lbl, code in OPEX_CATS:
        ws.cell(row=r, column=4, value=lbl).font = ARIAL
        ws.cell(row=r, column=7, value="[EUR m]").font = ARIAL
        cols = year_cols()
        for j, col in enumerate(cols):
            if j == 0:
                ws[f"{col}{r}"] = f"=Inputs!{col}{reg.sel[f'base:{code}']}"
            else:
                pc = cols[j-1]
                ws[f"{col}{r}"] = (f"={pc}{r}*(1+IFERROR({col}11/{pc}11-1,0)*"
                                   f"Inputs!{col}{reg.sel[f'el_pax:{code}']})")
            ws[f"{col}{r}"].font = ARIAL
        first = first or r; last = r; r += 1
    ws.cell(row=r + 1, column=4, value="Total operating costs").font = BOLD
    for col in year_cols():
        ws[f"{col}{r+1}"] = f"=SUM({col}{first}:{col}{last})"; ws[f"{col}{r+1}"].font = ARIAL
    return ("Opex", r + 1)

def build_summary_returns(wb, aero, nonaero, opex):
    ws = wb.create_sheet("Operations Summary")
    ws["A1"] = "=Control!B1"; ws["A1"].font = BOLD
    for j in range(NYEARS):
        ws.cell(row=6, column=FIRST_YCOL + j, value=START_YEAR + j).font = BOLD
    rows = [("Aeronautical revenue", f"='{aero[0]}'!{{c}}{aero[1]}"),
            ("Non-aeronautical revenue", f"='{nonaero[0]}'!{{c}}{nonaero[1]}"),
            ("Total revenue", "={c}8+{c}9"),
            ("Operating costs", f"='{opex[0]}'!{{c}}{opex[1]}"),
            ("EBITDA", "={c}10-{c}11")]
    for i, (lbl, f) in enumerate(rows):
        r = 8 + i
        ws.cell(row=r, column=4, value=lbl).font = BOLD if lbl in ("Total revenue","EBITDA") else ARIAL
        for col in year_cols():
            ws[f"{col}{r}"] = f.format(c=col); ws[f"{col}{r}"].font = ARIAL
    rt = wb.create_sheet("Returns")
    rt["A1"] = "=Control!B1"; rt["A1"].font = BOLD
    rt["D8"] = "Returns (v0 stub: entry/exit multiples on EBITDA, per blueprint section 4)"
    rt["D8"].font = ARIAL
    rt["D10"] = "EBITDA entry multiple [x]"; rt["H10"] = 10
    rt["D11"] = "EBITDA exit multiple [x]"; rt["H11"] = 10
    rt["D12"] = "Entry year"; rt["H12"] = START_YEAR + 1
    rt["D13"] = "Exit year"; rt["H13"] = START_YEAR + 10
    for r in range(10, 14): rt[f"D{r}"].font = ARIAL; rt[f"H{r}"].font = ARIAL
    return ws

def main(out_path):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    cover = build_cover(wb); ctrl = build_control(wb)
    inp = wb.create_sheet("Inputs")
    inp["A1"] = "=Control!B1"; inp["A1"].font = BOLD
    for j in range(NYEARS):
        inp.cell(row=6, column=FIRST_YCOL + j, value=START_YEAR + j).font = BOLD
    reg = Registry(); ctrl_row = 40
    for lbl, code, unit in TRAFFIC + ATMS:
        input_block(inp, ctrl, reg, lbl, code, unit, ctrl_row); ctrl_row += 1
    for lbl, code, drv in AERO_LINES:
        input_block(inp, ctrl, reg, f"{lbl} - unit rate", f"charge:{code}:{lbl}", "[EUR real]", ctrl_row); ctrl_row += 1
    for lbl, code in NONAERO_CATS:
        input_block(inp, ctrl, reg, f"{lbl} - base year revenue", f"baserev:{code}", "[EUR m]", ctrl_row); ctrl_row += 1
        input_block(inp, ctrl, reg, f"{lbl} - elasticity to pax growth", f"el_pax:{code}", "[x]", ctrl_row); ctrl_row += 1
        input_block(inp, ctrl, reg, f"{lbl} - elasticity to terminal size", f"el_term:{code}", "[x]", ctrl_row); ctrl_row += 1
        input_block(inp, ctrl, reg, f"{lbl} - uplift", f"uplift:{code}", "[%]", ctrl_row); ctrl_row += 1
    for lbl, code in OPEX_CATS:
        input_block(inp, ctrl, reg, f"{lbl} - base year", f"base:{code}", "[EUR m]", ctrl_row); ctrl_row += 1
        input_block(inp, ctrl, reg, f"{lbl} - elasticity to pax growth", f"el_pax:{code}", "[x]", ctrl_row); ctrl_row += 1
    aero = build_aero(wb, reg); nonaero = build_nonaero(wb, reg); opex = build_opex(wb, reg)
    build_summary_returns(wb, aero, nonaero, opex)
    wb.properties.creator = "Avia Solutions"
    wb.properties.lastModifiedBy = "Avia Solutions"
    wb.properties.title = "Avia Standard Airport Business-Plan Model - skeleton v0"
    wb.save(out_path)
    print("saved", out_path)

if __name__ == "__main__":
    # Output resolves from the data root, or is given on the command line. It previously
    # carried an absolute path from a working session, which wrote nowhere on any other host.
    import os, config
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        config.output_dir(), "Avia_Model_Skeleton_v0.xlsx")
    main(out)
